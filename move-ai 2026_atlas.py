import streamlit as st
import altair as alt
import folium
from folium.plugins import MiniMap, Fullscreen
import math, random, json, time, io, hashlib, urllib.parse, smtplib, sqlite3, os
from email.mime.text import MIMEText
from datetime import date, time as dtime, timedelta, datetime
import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup

try:
    import google.generativeai as genai
    GEMINI_AVAILABLE = True
except:
    GEMINI_AVAILABLE = False

try:
    from streamlit_folium import st_folium
    ST_FOLIUM_AVAILABLE = True
except:
    ST_FOLIUM_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except:
    FPDF_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except:
    OPENPYXL_AVAILABLE = False

try:
    import xlsxwriter
    XLSXWRITER_AVAILABLE = True
except:
    XLSXWRITER_AVAILABLE = False

st.set_page_config(page_title="ATLAS AI - Logistics Intelligence", layout="wide")

LANG = {
    "ko": {
        "app_title": "G-Navigator AI - Global Vehicle Logistics Navigator",
        "app_tagline": "**AI-Powered Supply Chain Intelligence | Real-time AIS Tracking | Risk Management**",
        "tab_dashboard": "대시보드",
        "tab_domestic": "국내운송 (Domestic)",
        "tab_intl": "국제운송 (International)",
        "domestic_header": "국내 운송 관리 시스템",
        "intl_header": "International Logistics Management",
        "dashboard_header": "ATLAS AI 홈 대시보드",

        "sidebar_domestic_settings": "국내운송 설정",
        "sidebar_intl_settings": "국제운송 설정",
        "sidebar_global_settings": "전역 설정 / API 연동",
        "cargo_type_label": "화물 종류",
        "dangerous_goods_label": "위험물",

        "fav_title": "즐겨찾기 경로 (영구 저장)",
        "fav_name_input": "현재 경로 저장 이름",
        "fav_save_btn": "현재 경로 즐겨찾기 저장",
        "fav_saved_msg": "저장 완료 (DB에 영구 저장됨)",
        "fav_delete_select": "삭제할 즐겨찾기",
        "fav_delete_btn": "선택 삭제",
        "fav_empty": "아직 저장된 즐겨찾기가 없습니다.",

        "report_title": "리포트 내보내기 및 공유",
        "report_pdf_btn": "PDF 다운로드",
        "report_pdf_missing": "PDF 내보내기는 `pip install fpdf2` 설치 후 사용 가능합니다.",
        "report_excel_btn": "Excel 다운로드",
        "report_excel_missing": "Excel 내보내기는 `pip install openpyxl` (또는 `xlsxwriter`) 설치 후 사용 가능합니다.",
        "report_email_header": "**이메일로 공유**",
        "report_email_to": "받는 사람 이메일 (선택)",
        "report_email_open_btn": "메일 클라이언트로 열기",
        "report_slack_header": "**Slack으로 공유**",
        "report_slack_msg_label": "Slack 메시지",
        "report_slack_send_btn": "Slack으로 전송",

        "compare_strategy_title": "전략 A/B/C 동시 비교",
        "compare_strategy_btn": "전체 전략 비교 실행",
        "compare_route_title": "경로 비교 (다른 목적지와 비교)",
        "compare_route_dest_label": "비교할 목적지",
        "compare_route_btn": "비교하기",

        "accept_reject_title": "운송 수락 판단 지원 (Call 수락 AI)",
        "accept_reject_price_label": "견적/청구 운임 (USD)",
        "accept_reject_no_sim": "먼저 시뮬레이션을 실행하면 견적가 대비 수익성을 분석할 수 있습니다.",
        "accept_reject_verdict_label": "판단",
        "accept_reject_margin_label": "예상 마진",
        "accept_reject_prob_label": "수익 확률",

        "reposition_title": "권역 이동 추천 (도착 후 재배치)",
        "social_benefit_title": "철도 vs 도로 사회·환경적 편익 비교",
        "rail_platform_title": "철도 물류 플랫폼",
        "rail_platform_cars_header": "**실시간 열차/빈 화차 정보 조회**",
        "rail_platform_ktx_header": "**KTX 특송 AI 접수**",
        "rail_platform_ktx_origin": "발송역",
        "rail_platform_ktx_dest": "도착역",
        "rail_platform_ktx_desc": "화물 설명",
        "rail_platform_ktx_submit": "KTX 특송 접수하기",

        "ai_assistant_title": "AI 어시스턴트 (Gemini)",
        "ai_question_label": "물류 관련 질문을 입력하세요",
        "ai_ask_btn": "질문하기",
        "ai_briefing_btn": "AI 리스크 브리핑 생성",

        "calculator_title": "리스크·비용 계산기",
        "forwarder_header": "**포워더 신뢰도 (샘플 데이터)**",

        "dashboard_shipments_title": "진행중인 배송 목록",
        "dashboard_shipments_empty": "아직 등록된 배송이 없습니다. 각 탭의 시뮬레이션 결과에서 '진행중인 배송으로 등록' 버튼으로 추가할 수 있어요.",
        "dashboard_status_change_label": "상태 변경할 배송",
        "dashboard_status_new_label": "새 상태",
        "dashboard_status_update_btn": "상태 업데이트",
        "dashboard_delete_label": "삭제할 배송",
        "dashboard_delete_btn": "배송 삭제",
        "dashboard_cost_trend_title": "비용 추이 (최근 시뮬레이션 이력)",
        "dashboard_cost_trend_empty": "아직 시뮬레이션 이력이 없습니다. 각 탭에서 시뮬레이션을 실행하면 여기에 누적됩니다.",
        "dashboard_visibility_title": "공급망 가시성 (전체 화물 현황)",
        "dashboard_pattern_title": "운행 패턴 분석 (암묵지 디지털 자산화)",
        "dashboard_pattern_empty": "아직 분석할 시뮬레이션 이력이 없습니다. 각 탭에서 시뮬레이션을 실행해보세요.",
        "dashboard_pattern_ai_btn": "AI 패턴 인사이트 생성",
        "dashboard_pattern_top_routes": "**가장 많이 조회된 경로 Top 5**",
        "dashboard_pattern_top_strategy": "**가장 많이 선택된 전략**",
        "dashboard_pattern_avg_risk": "**운송 유형별 평균 리스크**",

        "register_shipment_btn": "진행중인 배송으로 등록",
        "register_shipment_msg": "진행중인 배송으로 등록했습니다. 홈 대시보드에서 확인하세요.",

        "accept_reject_caption": "수익 확률 = 몬테카를로 시뮬레이션 1000회 중 견적가가 실제 비용보다 높았던 비율",
        "reposition_result_msg": "'{dest}' 도착 후에는 인근의 **{name}**(약 {dist:.0f}km, 혼잡도 {score}/100)로 이동하면 다음 화물을 더 빠르게 확보할 가능성이 높습니다.",
        "reposition_empty": "150km 이내 인근 터미널 데이터가 없습니다.",
        "social_benefit_co2_label": "CO2 절감",
        "social_benefit_air_label": "대기오염 개선(근사)",
        "social_benefit_accident_label": "사고 리스크 지수",
        "social_benefit_congestion_label": "도로혼잡 완화(근사)",
        "social_benefit_caption": "사회·환경적 편익은 거리 기반 근사치입니다. 실서비스 전환 시 온실가스종합정보센터/교통안전공단 통계 연동을 권장합니다.",
        "social_benefit_co2_delta": "도로 {road:.2f}→철도 {rail:.2f}",
        "social_benefit_accident_delta": "도로 {road} vs 철도 {rail}",
        "social_benefit_congestion_unit": "차·km 회피",
        "rail_col_terminal": "터미널", "rail_col_cars": "가용 화차 수", "rail_col_next": "다음 열차(분 후)",
        "ktx_placeholder": "예: 반도체 긴급부품 20kg",
        "ktx_success_msg": "KTX 특송 접수 완료! ({desc}) - 홈 대시보드에서 확인하세요.",
        "ktx_no_desc": "화물 설명 없음",
        "label_strategy": "전략", "label_count": "횟수", "label_route_type": "운송유형",
        "label_avg_risk": "평균 리스크", "label_destination": "목적지",
        "pattern_ai_spinner": "패턴 분석 중...",
        "pattern_ai_no_response": "Gemini 응답을 가져오지 못했습니다.",
        "pattern_ai_hint": "Gemini API 키를 등록하면 AI가 패턴에서 인사이트를 요약해줍니다.",
        "delay_buffer_warning": "예상 소요시간의 변동성이 높습니다 (표준편차 {std:.1f}일). 리드타임에 최소 **{buf}일**의 버퍼를 추가하는 것을 권장합니다.",

        "dashboard_kpi_shipments": "진행중인 배송 건수",
        "dashboard_kpi_avg_risk": "평균 리스크 스코어",
        "dashboard_kpi_total_cost": "총 예상비용",
        "dashboard_kpi_sim_count": "누적 시뮬레이션 횟수",
    },
    "en": {
        "app_title": "G-Navigator AI - Global Vehicle Logistics Navigator",
        "app_tagline": "**AI-Powered Supply Chain Intelligence | Real-time AIS Tracking | Risk Management**",
        "tab_dashboard": "Dashboard",
        "tab_domestic": "Domestic Logistics",
        "tab_intl": "International Logistics",
        "domestic_header": "Domestic Logistics Management System",
        "intl_header": "International Logistics Management",
        "dashboard_header": "ATLAS AI Home Dashboard",

        "sidebar_domestic_settings": "Domestic Settings",
        "sidebar_intl_settings": "International Settings",
        "sidebar_global_settings": "Global Settings / API Integration",
        "cargo_type_label": "Cargo Type",
        "dangerous_goods_label": "Dangerous Goods",

        "fav_title": "Favorite Routes (Saved Permanently)",
        "fav_name_input": "Name for current route",
        "fav_save_btn": "Save current route to favorites",
        "fav_saved_msg": "Saved (persisted to database)",
        "fav_delete_select": "Favorite to delete",
        "fav_delete_btn": "Delete selected",
        "fav_empty": "No favorites saved yet.",

        "report_title": "Export Report & Share",
        "report_pdf_btn": "Download PDF",
        "report_pdf_missing": "PDF export requires `pip install fpdf2`.",
        "report_excel_btn": "Download Excel",
        "report_excel_missing": "Excel export requires `pip install openpyxl` (or `xlsxwriter`).",
        "report_email_header": "**Share via Email**",
        "report_email_to": "Recipient email (optional)",
        "report_email_open_btn": "Open in mail client",
        "report_slack_header": "**Share via Slack**",
        "report_slack_msg_label": "Slack message",
        "report_slack_send_btn": "Send to Slack",

        "compare_strategy_title": "Compare Strategies A/B/C",
        "compare_strategy_btn": "Run full strategy comparison",
        "compare_route_title": "Compare Routes (vs. another destination)",
        "compare_route_dest_label": "Destination to compare",
        "compare_route_btn": "Compare",

        "accept_reject_title": "Shipment Accept/Reject Advisor (Call AI)",
        "accept_reject_price_label": "Quoted freight rate (USD)",
        "accept_reject_no_sim": "Run a simulation first to analyze profitability against the quoted price.",
        "accept_reject_verdict_label": "Verdict",
        "accept_reject_margin_label": "Expected Margin",
        "accept_reject_prob_label": "Profit Probability",

        "reposition_title": "Repositioning Suggestion (after arrival)",
        "social_benefit_title": "Rail vs Road Social & Environmental Benefit Comparison",
        "rail_platform_title": "Rail Logistics Platform",
        "rail_platform_cars_header": "**Real-time Train / Empty Freight Car Info**",
        "rail_platform_ktx_header": "**KTX Express AI Booking**",
        "rail_platform_ktx_origin": "Departure Station",
        "rail_platform_ktx_dest": "Arrival Station",
        "rail_platform_ktx_desc": "Cargo Description",
        "rail_platform_ktx_submit": "Submit KTX Express Booking",

        "ai_assistant_title": "AI Assistant (Gemini)",
        "ai_question_label": "Ask a logistics-related question",
        "ai_ask_btn": "Ask",
        "ai_briefing_btn": "Generate AI Risk Briefing",

        "calculator_title": "Risk & Cost Calculators",
        "forwarder_header": "**Forwarder Reputation (sample data)**",

        "dashboard_shipments_title": "Active Shipments",
        "dashboard_shipments_empty": "No shipments registered yet. Use the 'Register as active shipment' button in either tab's simulation results.",
        "dashboard_status_change_label": "Shipment to update",
        "dashboard_status_new_label": "New status",
        "dashboard_status_update_btn": "Update status",
        "dashboard_delete_label": "Shipment to delete",
        "dashboard_delete_btn": "Delete shipment",
        "dashboard_cost_trend_title": "Cost Trend (recent simulation history)",
        "dashboard_cost_trend_empty": "No simulation history yet. Run a simulation in either tab to start accumulating data.",
        "dashboard_visibility_title": "Supply Chain Visibility (all shipments)",
        "dashboard_pattern_title": "Operating Pattern Analysis (Tacit Knowledge Digitization)",
        "dashboard_pattern_empty": "No simulation history to analyze yet. Try running a simulation in either tab.",
        "dashboard_pattern_ai_btn": "Generate AI Pattern Insight",
        "dashboard_pattern_top_routes": "**Top 5 Most Viewed Routes**",
        "dashboard_pattern_top_strategy": "**Most Selected Strategy**",
        "dashboard_pattern_avg_risk": "**Average Risk by Route Type**",

        "register_shipment_btn": "Register as active shipment",
        "register_shipment_msg": "Registered as an active shipment. Check the home dashboard.",

        "accept_reject_caption": "Profit probability = share of 1,000 Monte Carlo runs where the quoted price exceeded the actual cost",
        "reposition_result_msg": "After arriving at '{dest}', moving to nearby **{name}** (~{dist:.0f}km, congestion {score}/100) increases the chance of securing the next load faster.",
        "reposition_empty": "No nearby terminal data within 150km.",
        "social_benefit_co2_label": "CO2 Reduction",
        "social_benefit_air_label": "Air Pollution Improvement (approx.)",
        "social_benefit_accident_label": "Accident Risk Index",
        "social_benefit_congestion_label": "Congestion Relief (approx.)",
        "social_benefit_caption": "Social/environmental benefits are distance-based approximations. For production use, integrate with GIR/KoROAD statistics.",
        "social_benefit_co2_delta": "Road {road:.2f}→Rail {rail:.2f}",
        "social_benefit_accident_delta": "Road {road} vs Rail {rail}",
        "social_benefit_congestion_unit": "vehicle·km avoided",
        "rail_col_terminal": "Terminal", "rail_col_cars": "Available Cars", "rail_col_next": "Next Train (min)",
        "ktx_placeholder": "e.g. Urgent semiconductor parts 20kg",
        "ktx_success_msg": "KTX Express booking complete! ({desc}) - Check the home dashboard.",
        "ktx_no_desc": "No description",
        "label_strategy": "Strategy", "label_count": "Count", "label_route_type": "Route Type",
        "label_avg_risk": "Avg Risk", "label_destination": "Destination",
        "pattern_ai_spinner": "Analyzing patterns...",
        "pattern_ai_no_response": "Could not get a response from Gemini.",
        "pattern_ai_hint": "Register a Gemini API key to get AI-generated pattern insights.",
        "delay_buffer_warning": "Estimated delivery time shows high variability (std dev {std:.1f} days). We recommend adding at least **{buf} days** of lead-time buffer.",

        "dashboard_kpi_shipments": "Active Shipments",
        "dashboard_kpi_avg_risk": "Avg Risk Score",
        "dashboard_kpi_total_cost": "Total Expected Cost",
        "dashboard_kpi_sim_count": "Total Simulations",
    },
}

def t(key):
    lang = st.session_state.get("nav_language", "ko") if hasattr(st, "session_state") else "ko"
    return LANG.get(lang, LANG["ko"]).get(key, LANG["ko"].get(key, key))

st.title(t("app_title"))
st.markdown(t("app_tagline"))

DOMESTIC_TERMINALS = {
    "부산항": [129.0430, 35.0979, "Port", "Mega Hub"],
    "인천항": [126.6297, 37.3559, "Port", "Major Hub"],
    "광양항": [127.7347, 34.9405, "Port", "Major Hub"],
    "울산항": [129.3893, 35.0427, "Port", "Major Hub"],
    "목포항": [126.3823, 34.7712, "Port", "Regional Hub"],
    "군산항": [126.7186, 35.9772, "Port", "Regional Hub"],
    "여수항": [127.7416, 34.7604, "Port", "Regional Hub"],
    "동해항": [129.1196, 37.5203, "Port", "Regional Hub"],
    "속초항": [128.5934, 38.2038, "Port", "Regional Hub"],

    "인천공항": [126.4406, 37.4602, "Airport", "Mega Hub"],
    "김포공항": [126.8010, 37.5580, "Airport", "Major Hub"],
    "부산공항": [129.0676, 35.1792, "Airport", "Major Hub"],
    "대구공항": [128.6556, 35.8949, "Airport", "Regional Hub"],
    "광주공항": [126.8048, 35.1151, "Airport", "Regional Hub"],

    "서울": [126.9780, 37.5665, "Terminal", "Major Hub"],
    "대전": [127.4245, 36.3504, "Terminal", "Regional Hub"],
    "대구": [128.5955, 35.8714, "Terminal", "Regional Hub"],
    "시흥IC": [126.8100, 37.3719, "Terminal", "Regional Hub"],
    "평택": [127.0073, 37.0274, "Terminal", "Regional Hub"],
    "이천": [127.2280, 37.1943, "Terminal", "Regional Hub"],
    "충주": [127.9346, 36.9915, "Terminal", "Regional Hub"],
    "김천": [128.1146, 35.9928, "Terminal", "Regional Hub"],
}

DOMESTIC_TERMINAL_NAMES_EN = {
    "부산항": "Busan Port", "인천항": "Incheon Port", "광양항": "Gwangyang Port", "울산항": "Ulsan Port",
    "목포항": "Mokpo Port", "군산항": "Gunsan Port", "여수항": "Yeosu Port", "동해항": "Donghae Port", "속초항": "Sokcho Port",
    "인천공항": "Incheon Airport", "김포공항": "Gimpo Airport", "부산공항": "Busan Airport",
    "대구공항": "Daegu Airport", "광주공항": "Gwangju Airport",
    "서울": "Seoul", "대전": "Daejeon", "대구": "Daegu", "시흥IC": "Siheung IC",
    "평택": "Pyeongtaek", "이천": "Icheon", "충주": "Chungju", "김천": "Gimcheon",
}

def display_terminal_name(name):
    if st.session_state.get("nav_language", "ko") == "en":
        return DOMESTIC_TERMINAL_NAMES_EN.get(name, name)
    return name

PORTS_DB = {
    "Busan (부산, Korea)": [129.0755, 35.0982, "South Korea", "Mega Hub"],
    "Incheon (인천, Korea)": [126.6270, 37.3592, "South Korea", "Major Hub"],
    "Shanghai (상하이, China)": [121.4737, 31.2304, "China", "Mega Hub"],
    "Tianjin (천진, China)": [117.7010, 39.0842, "China", "Major Hub"],
    "Ningbo (닝보, China)": [121.5650, 29.8683, "China", "Major Hub"],
    "Hong Kong (홍콩)": [114.1628, 22.3193, "Hong Kong", "Mega Hub"],
    "Singapore (싱가포르)": [103.8509, 1.3521, "Singapore", "Transshipment"],
    "Kaohsiung (고雄, Taiwan)": [120.2708, 22.6117, "Taiwan", "Major Hub"],
    "Port Klang (말레이시아)": [101.5230, 3.3256, "Malaysia", "Regional Hub"],
    "Bangkok (방콕, Thailand)": [100.5881, 13.7249, "Thailand", "Regional Hub"],
    "Tokyo (도쿄, Japan)": [139.7673, 35.4437, "Japan", "Major Hub"],
    "Yokohama (요코하마, Japan)": [139.6380, 35.4437, "Japan", "Major Hub"],
    "Kobe (고베, Japan)": [135.1955, 34.6901, "Japan", "Regional Hub"],
    "Ho Chi Minh (호치민, Vietnam)": [106.6869, 10.7769, "Vietnam", "Regional Hub"],

    "Dubai (두바이, UAE)": [55.2708, 25.2048, "UAE", "Mega Hub"],
    "Jebel Ali (제벨알리, UAE)": [55.1170, 24.9774, "UAE", "Mega Hub"],
    "Karachi (카라치, Pakistan)": [67.0099, 24.7569, "Pakistan", "Major Hub"],
    "Mumbai (뭄바이, India)": [72.8479, 19.0176, "India", "Major Hub"],

    "Rotterdam (로테르담)": [4.2917, 51.9225, "Netherlands", "Mega Hub"],
    "Hamburg (함부르크)": [10.0086, 53.3456, "Germany", "Mega Hub"],
    "Antwerp (앤트워프)": [4.4699, 51.3397, "Belgium", "Mega Hub"],

    "Los Angeles (LA)": [-118.2437, 33.7490, "USA", "Mega Hub"],
    "Long Beach (롱비치)": [-118.1937, 33.7381, "USA", "Mega Hub"],
    "New York (뉴욕)": [-74.0060, 40.7128, "USA", "Mega Hub"],
}

GEOPOLITICAL_RISKS = {
    "Strait of Hormuz (호르무즈 해협)": {
        "coords": [[25.5, 56.5], [25.5, 57.0], [26.0, 57.0], [26.0, 56.5]],
        "risk_level": "High",
        "impact": 1.35,
        "description": "중동 석유 운송의 중요 경로, 지정학적 긴장 시 운임 급등"
    },
    "Suez Canal (수에즈 운하)": {
        "coords": [[31.5, 31.0], [31.5, 32.0], [32.0, 32.0], [32.0, 31.0]],
        "risk_level": "Medium",
        "impact": 1.20,
        "description": "유럽-아시아 최단 경로, 정치적 불안정 시 폐쇄 위험"
    },
    "Red Sea (홍해)": {
        "coords": [[12.0, 41.0], [12.0, 44.0], [16.0, 44.0], [16.0, 41.0]],
        "risk_level": "High",
        "impact": 1.30,
        "description": "후티 무장단 활동 증가, 선박 피해 위험"
    },
    "South China Sea (남중국해)": {
        "coords": [[4.0, 100.0], [4.0, 120.0], [12.0, 120.0], [12.0, 100.0]],
        "risk_level": "Medium",
        "impact": 1.15,
        "description": "영토 분쟁 지역, 해운 자유도 제약 가능성"
    },
}

TRANSPORT_MODES = {
    "AIR": {"cost_multiplier": 3.5, "time_multiplier": 0.15, "risk": 35, "co2": 1.8, "capacity": "Limited"},
    "SEA": {"cost_multiplier": 1.0, "time_multiplier": 1.0, "risk": 55, "co2": 0.2, "capacity": "Very High"},
    "RAIL": {"cost_multiplier": 0.8, "time_multiplier": 0.7, "risk": 40, "co2": 0.5, "capacity": "High"},
    "ROAD": {"cost_multiplier": 1.2, "time_multiplier": 0.6, "risk": 45, "co2": 1.0, "capacity": "Medium"},
}

BULK_CARGO_TYPES = [
    "Coal", "Iron Ore", "Grain", "Bauxite", "Phosphate",
    "Sulfur", "Fertilizer", "Wood Chips", "Cement", "Custom"
]

DOOR_TO_DOOR_LASTMILE_DAYS = 2.5
DOOR_TO_DOOR_LASTMILE_COST_RATE = 0.12
DOMESTIC_TRUCK_SPEED_KMH = 60

LCL_CONSOLIDATION_EXTRA_DAYS = 3.0
LCL_COST_MULTIPLIER = 1.15
LCL_RATE_PER_KG_USD = 0.8

INCOTERMS_CLEARANCE_DAYS = {
    "EXW": {"export_days": 3, "import_days": 4, "note": "구매자가 수출/수입 통관 모두 책임 (실무상 셀러 협조 필요)"},
    "FCA": {"export_days": 2, "import_days": 4, "note": "판매자 수출통관, 구매자 수입통관"},
    "FAS": {"export_days": 2, "import_days": 4, "note": "판매자 수출통관, 구매자 수입통관"},
    "FOB": {"export_days": 2, "import_days": 4, "note": "판매자 수출통관, 구매자 수입통관"},
    "CFR": {"export_days": 2, "import_days": 4, "note": "판매자 수출통관, 구매자 수입통관"},
    "CIF": {"export_days": 2, "import_days": 4, "note": "판매자 수출통관+보험, 구매자 수입통관"},
    "CPT": {"export_days": 2, "import_days": 3, "note": "판매자 수출통관, 구매자 수입통관"},
    "CIP": {"export_days": 2, "import_days": 3, "note": "판매자 수출통관+보험, 구매자 수입통관"},
    "DAP": {"export_days": 2, "import_days": 3, "note": "판매자가 목적지까지 책임, 수입통관은 구매자"},
    "DPU": {"export_days": 2, "import_days": 3, "note": "판매자가 양하까지 책임"},
    "DDP": {"export_days": 1, "import_days": 1, "note": "판매자가 수출/수입 통관 모두 책임 (구매자 부담 최소)"},
}

CARGO_CLEARANCE_EXTRA_DAYS = {
    "위험물": 2, "Dangerous Goods": 2,
    "식품": 1, "Batteries": 2, "Semiconductors": 1,
}

EU_CARBON_PRICE_EUR_PER_TON = 90.0

FORWARDER_REPUTATION = {
    "HMM": {"on_time_rate": 92.5, "incident_rate": 1.2, "avg_rating": 4.4},
    "Maersk": {"on_time_rate": 90.1, "incident_rate": 1.5, "avg_rating": 4.3},
    "MSC": {"on_time_rate": 88.7, "incident_rate": 1.8, "avg_rating": 4.1},
    "CMA CGM": {"on_time_rate": 89.4, "incident_rate": 1.6, "avg_rating": 4.2},
    "COSCO": {"on_time_rate": 86.9, "incident_rate": 2.1, "avg_rating": 3.9},
    "ONE": {"on_time_rate": 91.0, "incident_rate": 1.3, "avg_rating": 4.3},
    "Evergreen": {"on_time_rate": 87.8, "incident_rate": 1.9, "avg_rating": 4.0},
    "SM Line": {"on_time_rate": 90.6, "incident_rate": 1.4, "avg_rating": 4.2},
}

if 'selected_strategy' not in st.session_state:
    st.session_state.selected_strategy = "STRATEGY_B"

if 'run_simulation' not in st.session_state:
    st.session_state.run_simulation = False

if 'logistics_insights' not in st.session_state:
    st.session_state.logistics_insights = None

if 'current_tab' not in st.session_state:
    st.session_state.current_tab = 0

if 'ais_vessels_domestic' not in st.session_state:
    st.session_state.ais_vessels_domestic = None

if 'ais_vessels_intl' not in st.session_state:
    st.session_state.ais_vessels_intl = None

if 'currency' not in st.session_state:
    st.session_state.currency = "USD"
if 'exchange_rates' not in st.session_state:
    st.session_state.exchange_rates = None
if 'dark_mode' not in st.session_state:
    st.session_state.dark_mode = False
if 'compact_mode' not in st.session_state:
    st.session_state.compact_mode = False
if 'ai_chat_history' not in st.session_state:
    st.session_state.ai_chat_history = []
if 'ai_risk_briefing' not in st.session_state:
    st.session_state.ai_risk_briefing = None
if 'typhoon_data' not in st.session_state:
    st.session_state.typhoon_data = None
if 'port_congestion_cache' not in st.session_state:
    st.session_state.port_congestion_cache = {}
if 'map_zoom_leg_domestic' not in st.session_state:
    st.session_state.map_zoom_leg_domestic = "전체 경로"
if 'map_zoom_leg_intl' not in st.session_state:
    st.session_state.map_zoom_leg_intl = "전체 경로"
if 'aishub_username' not in st.session_state:
    st.session_state.aishub_username = ""
if 'kma_api_key' not in st.session_state:
    st.session_state.kma_api_key = ""
if 'gemini_api_key' not in st.session_state:
    st.session_state.gemini_api_key = ""
if 'slack_webhook' not in st.session_state:
    st.session_state.slack_webhook = ""
if 'customs_api_key' not in st.session_state:
    st.session_state.customs_api_key = ""
if 'nav_language' not in st.session_state:
    st.session_state.nav_language = "ko"
if 'risk_alert_threshold' not in st.session_state:
    st.session_state.risk_alert_threshold = 70
if 'auto_slack_alert' not in st.session_state:
    st.session_state.auto_slack_alert = False

def haversine_km(lon1, lat1, lon2, lat2):
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    return 2*R*math.asin(math.sqrt(a))

def create_ai_route_with_waypoints(origin_coord, waypoints, dest_coord, num_points=15, offset_km=20):
    all_points = [origin_coord] + waypoints + [dest_coord]
    route_points = []

    for i in range(len(all_points) - 1):
        lon1, lat1 = all_points[i]
        lon2, lat2 = all_points[i + 1]

        for j in range(num_points):
            t = j / (num_points - 1)

            lon = lon1 + (lon2 - lon1) * t
            lat = lat1 + (lat2 - lat1) * t

            gauss = math.exp(-((t - 0.5) ** 2) / (2 * 0.25 ** 2))
            offset = offset_km * gauss * 0.5

            dx = lon2 - lon1
            dy = lat2 - lat1
            dist = math.sqrt(dx**2 + dy**2)

            if dist > 0:
                perp_x = -dy / dist
                perp_y = dx / dist

                deg_offset_lon = offset / 111.0
                deg_offset_lat = offset / 110.6

                lon_offset = lon + perp_x * deg_offset_lon
                lat_offset = lat + perp_y * deg_offset_lat
            else:
                lon_offset = lon
                lat_offset = lat

            route_points.append([lat_offset, lon_offset])

    return route_points

def apply_geopolitical_risk(route_distance, selected_risks):
    risk_multiplier = 1.0
    risk_description = []

    for risk_name in selected_risks:
        if risk_name in GEOPOLITICAL_RISKS:
            multiplier = GEOPOLITICAL_RISKS[risk_name]["impact"]
            risk_multiplier *= multiplier
            risk_description.append(f"{risk_name}: +{int((multiplier-1)*100)}%")

    adjusted_distance = route_distance * (risk_multiplier ** 0.3)
    return adjusted_distance, risk_multiplier, risk_description

def fetch_logistics_news():
    try:
        url = "https://www.kita.net/shippers/board/newsList.do"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(url, headers=headers, timeout=5)
        response.encoding = 'utf-8'

        soup = BeautifulSoup(response.content, 'html.parser')

        news_items = []
        articles = soup.find_all('tr')[:5]

        for article in articles:
            cols = article.find_all('td')
            if len(cols) >= 2:
                title = cols[1].get_text(strip=True)
                if title:
                    news_items.append(title[:80])

        return news_items if news_items else ["최신 물류 뉴스를 불러올 수 없습니다"]
    except Exception as e:
        return [f"뉴스 로드 오류: {str(e)[:50]}"]

def render_ai_dashboard():
    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric("Global Port Congestion", "3.4M TEU", "+7.5 days")

    with col2:
        st.metric("Logistics Cost Index", "125", "+25%")

    with col3:
        st.metric("Material Cost Rise", "+12.3%", "YoY")

    with col4:
        st.metric("Geopolitical Risk", "CRITICAL", "Red Sea & Middle East")

    st.markdown("---")
    st.subheader("Real-time Global Maritime Intelligence (MarineTraffic AIS)")
    col_left, col_right = st.columns([3, 1])

    with col_left:
        st.write("**Active Vessels: 150+** | **Ports: 50+** | **Risk Zones** | ━━ **Routes**")
        st.info("**AI Learning:** 최신 물류 뉴스와 해운 동향을 실시간 분석 중...")

    with col_right:
        st.write("**AI Status**")
        st.success("Learning Active")
        st.write("Updates: Real-time")

def get_vessel_color(vessel_type):
    color_map = {
        "Container": "#FF6B6B",
        "Bulk Carrier": "#4ECDC4",
        "Tanker": "#FFE66D",
        "General Cargo": "#95E1D3",
        "RoRo": "#C7CEEA",
        "Multipurpose": "#B5EAD7",
    }
    return color_map.get(vessel_type, "#95A5A6")

GLOBAL_SHIPPING_LANES = [
    {"path": [[128.5, 34.0], [125.0, 32.0], [122.5, 30.5], [121.8, 30.0]],
     "lon_jitter": (-0.4, 0.4), "lat_jitter": (-0.3, 0.3), "name_prefix": "EAST CHINA SEA"},
    {"path": [[114.0, 20.0], [112.0, 15.0], [110.0, 10.0], [107.0, 5.0], [104.0, 2.0]],
     "lon_jitter": (-0.5, 0.5), "lat_jitter": (-0.5, 0.5), "name_prefix": "SOUTH CHINA SEA"},
    {"path": [[103.8, 1.3], [100.5, 3.0], [97.5, 4.5], [95.0, 6.0]],
     "lon_jitter": (-0.1, 0.1), "lat_jitter": (-0.08, 0.08), "name_prefix": "MALACCA STRAIT"},
    {"path": [[80.0, 6.0], [75.0, 8.0], [68.0, 10.0], [60.0, 11.5], [52.0, 13.0]],
     "lon_jitter": (-0.8, 0.8), "lat_jitter": (-0.6, 0.6), "name_prefix": "INDIAN OCEAN"},
    {"path": [[43.3, 12.6], [40.0, 17.0], [37.0, 22.0], [34.5, 27.5]],
     "lon_jitter": (-0.2, 0.2), "lat_jitter": (-0.15, 0.15), "name_prefix": "RED SEA"},
    {"path": [[32.55, 31.25], [32.5, 30.6], [32.35, 29.95]],
     "lon_jitter": (-0.03, 0.03), "lat_jitter": (-0.03, 0.03), "name_prefix": "SUEZ CANAL"},
    {"path": [[30.0, 33.0], [20.0, 34.5], [10.0, 36.0], [0.0, 36.5], [-5.5, 36.0]],
     "lon_jitter": (-0.5, 0.5), "lat_jitter": (-0.4, 0.4), "name_prefix": "MEDITERRANEAN"},
    {"path": [[-6.0, 36.0], [-20.0, 36.0], [-40.0, 38.0], [-60.0, 39.0], [-73.5, 40.3]],
     "lon_jitter": (-1.5, 1.5), "lat_jitter": (-1.2, 1.2), "name_prefix": "ATLANTIC"},
    {"path": [[-79.5, 25.0], [-78.0, 30.0], [-75.5, 35.0], [-73.0, 40.0], [-70.5, 42.5]],
     "lon_jitter": (0.2, 0.8), "lat_jitter": (-0.3, 0.3), "name_prefix": "US EAST COAST"},
    {"path": [[-117.5, 32.7], [-119.5, 34.5], [-121.5, 37.0], [-123.5, 40.0], [-124.5, 44.0], [-124.5, 47.5]],
     "lon_jitter": (-0.8, -0.2), "lat_jitter": (-0.3, 0.3), "name_prefix": "US WEST COAST"},
    {"path": [[56.3, 26.6], [55.0, 26.0], [53.5, 25.5], [52.0, 25.3]],
     "lon_jitter": (-0.15, 0.15), "lat_jitter": (-0.1, 0.1), "name_prefix": "PERSIAN GULF"},
    {"path": [[129.5, 34.8], [129.8, 34.3], [130.2, 33.8]],
     "lon_jitter": (-0.15, 0.15), "lat_jitter": (-0.1, 0.1), "name_prefix": "KOREA STRAIT"},
]

DOMESTIC_SHIPPING_LANES = [
    {"path": [[125.3, 37.2], [125.1, 36.3], [125.3, 35.4], [125.6, 34.7]],
     "lon_jitter": (-0.3, 0.1), "lat_jitter": (-0.2, 0.2), "name_prefix": "WEST SEA"},
    {"path": [[126.3, 34.5], [127.0, 34.35], [127.7, 34.45], [128.5, 34.65], [129.05, 34.85]],
     "lon_jitter": (-0.2, 0.2), "lat_jitter": (-0.25, -0.02), "name_prefix": "SOUTH SEA"},
    {"path": [[129.4, 35.05], [129.7, 35.85], [129.6, 36.9], [129.35, 37.6], [129.15, 38.25]],
     "lon_jitter": (0.05, 0.4), "lat_jitter": (-0.25, 0.25), "name_prefix": "EAST SEA"},
]

def _generate_vessels_along_lanes(lanes, num_vessels, id_start, ship_names, flag_choices):
    vessels = []
    vessel_types = ["Container", "Bulk Carrier", "Tanker", "General Cargo", "RoRo", "Multipurpose"]
    vessel_id = id_start
    per_lane = max(1, int(num_vessels / len(lanes)))

    for lane in lanes:
        path = lane["path"]
        lon_j = lane["lon_jitter"]
        lat_j = lane["lat_jitter"]

        for _ in range(per_lane):
            seg_idx = random.randint(0, len(path) - 2)
            p1, p2 = path[seg_idx], path[seg_idx + 1]
            t = random.random()

            lon = p1[0] + (p2[0] - p1[0]) * t + random.uniform(lon_j[0], lon_j[1])
            lat = p1[1] + (p2[1] - p1[1]) * t + random.uniform(lat_j[0], lat_j[1])

            dx, dy = p2[0] - p1[0], p2[1] - p1[1]
            heading = (math.degrees(math.atan2(dx, dy)) + random.uniform(-15, 15)) % 360

            speed = random.uniform(8, 20)
            vessel_type = random.choice(vessel_types)
            ship_name = random.choice(ship_names)

            vessels.append({
                "mmsi": id_start * 1000 + vessel_id,
                "name": f"{ship_name} {vessel_id % 100}",
                "type": vessel_type,
                "lat": lat,
                "lon": lon,
                "heading": heading,
                "speed": speed,
                "status": "Underway",
                "flag": random.choice(flag_choices),
                "length": random.randint(80, 400),
                "beam": random.randint(15, 60),
            })
            vessel_id += 1

    return vessels

def generate_realistic_ais_vessels(num_vessels=150):
    ship_names = ["Ever Given", "MSC Gulsun", "COSCO", "Maersk", "CMA CGM", "Evergreen",
                  "ONE", "Hapag", "HMM", "Yang Ming", "ZIM", "Pacific", "Atlantic", "Indian"]
    flags = ["Panama", "Liberia", "Marshall Islands", "Hong Kong", "Singapore"]
    return _generate_vessels_along_lanes(GLOBAL_SHIPPING_LANES, num_vessels, 9, ship_names, flags)

def generate_realistic_ais_vessels_domestic(num_vessels=60):
    ship_names = ["HMM", "SM Line", "Pan Ocean", "KMTC", "Sinokor", "Heung-A", "Dong Young"]
    flags = ["South Korea"]
    return _generate_vessels_along_lanes(DOMESTIC_SHIPPING_LANES, num_vessels, 4, ship_names, flags)

def render_ais_vessels_on_map(m, vessels):
    for vessel in vessels:
        lat = vessel["lat"]
        lon = vessel["lon"]
        name = vessel["name"]
        mmsi = vessel["mmsi"]
        vessel_type = vessel["type"]
        heading = vessel["heading"]
        speed = vessel["speed"]
        status = vessel["status"]
        flag = vessel["flag"]

        vessel_color = get_vessel_color(vessel_type)

        popup_html = f"""
        <div style="font-family: Arial; font-size: 11px; width: 220px; padding: 8px;">
            <b style="color: {vessel_color}; font-size: 12px;">{name}</b>
            <hr style="margin: 4px 0; border: none; border-top: 1px solid #ccc;">
            <table style="width: 100%;">
                <tr><td><b>MMSI:</b></td><td>{mmsi}</td></tr>
                <tr><td><b>Type:</b></td><td>{vessel_type}</td></tr>
                <tr><td><b>Flag:</b></td><td>{flag}</td></tr>
                <tr><td><b>Speed:</b></td><td><b style="color: green;">{speed:.1f} kn</b></td></tr>
                <tr><td><b>Heading:</b></td><td>{heading:.0f}°</td></tr>
                <tr><td><b>Status:</b></td><td>{status}</td></tr>
                <tr><td><b>Pos:</b></td><td>{lat:.2f}°N / {lon:.2f}°E</td></tr>
            </table>
        </div>
        """

        folium.RegularPolygonMarker(
            location=[lat, lon],
            fill_color=vessel_color,
            number_of_sides=3,
            radius=5,
            rotation=heading,
            popup=folium.Popup(popup_html, max_width=250),
            tooltip=f"{name} ({speed:.0f}kn)",
            color=vessel_color,
            weight=1,
            fill_opacity=0.9
        ).add_to(m)

def render_risk_zones_circle(m, selected_risks):
    for risk_name, risk_info in GEOPOLITICAL_RISKS.items():
        coords = risk_info["coords"]
        center_lat = sum(c[0] for c in coords) / len(coords)
        center_lon = sum(c[1] for c in coords) / len(coords)

        risk_level = risk_info["risk_level"]
        radius_m = 350000 if risk_level == "High" else 250000
        is_selected = risk_name in selected_risks

        color = "#FF0000"
        fill_opacity = 0.28 if is_selected else 0.12
        weight = 2 if is_selected else 1
        status_text = "선택됨 - 비용에 반영중" if is_selected else "선택 시 운임/기간에 자동 반영됩니다"

        folium.Circle(
            location=[center_lat, center_lon],
            radius=radius_m,
            color=color,
            fill=True,
            fillColor=color,
            fillOpacity=fill_opacity,
            weight=weight,
            popup=f"<b>{risk_name}</b><br>Level: {risk_level}<br>{risk_info['description']}<br><i>{status_text}</i>",
            tooltip=f"{risk_name}" + (" (적용중)" if is_selected else "")
        ).add_to(m)

def bar_chart_horizontal_labels(series, category_name="Category", value_name="Count"):
    df = series.reset_index()
    df.columns = [category_name, value_name]
    chart = (
        alt.Chart(df)
        .mark_bar(color="#0099FF")
        .encode(
            x=alt.X(f"{category_name}:N", sort=None, axis=alt.Axis(labelAngle=0, title=None)),
            y=alt.Y(f"{value_name}:Q", title=value_name),
            tooltip=[category_name, value_name],
        )
        .properties(height=300)
    )
    st.altair_chart(chart, use_container_width=True)

def render_ais_fleet_statistics(vessels, coverage_label="Global"):
    st.subheader("Real-time AIS Fleet Statistics")

    vessel_df = pd.DataFrame(vessels)

    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        st.metric("Total Vessels", len(vessels))

    with col2:
        st.metric("Avg Speed", f"{vessel_df['speed'].mean():.1f} kn")

    with col3:
        st.metric("Max Speed", f"{vessel_df['speed'].max():.1f} kn")

    with col4:
        st.metric("Types", vessel_df['type'].nunique())

    with col5:
        st.metric("Coverage", coverage_label)

    chart_col1, chart_col2 = st.columns(2)

    with chart_col1:
        st.write("**Vessel Type Distribution**")
        bar_chart_horizontal_labels(vessel_df["type"].value_counts(), "Type", "Count")

    with chart_col2:
        st.write("**Speed Distribution by Range (knots)**")
        speed_ranges = []
        for speed in vessel_df["speed"]:
            if speed < 10:
                speed_ranges.append("8-10 kn")
            elif speed < 14:
                speed_ranges.append("10-14 kn")
            elif speed < 18:
                speed_ranges.append("14-18 kn")
            else:
                speed_ranges.append("18-25 kn")
        bar_chart_horizontal_labels(pd.Series(speed_ranges).value_counts(), "Speed Range", "Count")

def render_simulation_charts(sim_result):
    col_chart1, col_chart2, col_chart3 = st.columns(3)

    with col_chart1:
        st.write("**Cost Distribution**")
        st.line_chart(pd.DataFrame({"Cost": sorted(sim_result['costs'])[:100]}))

    with col_chart2:
        st.write("**Time Distribution**")
        st.line_chart(pd.DataFrame({"Days": sorted(sim_result['times'])[:100]}))

    with col_chart3:
        st.write("**Risk Distribution**")
        st.line_chart(pd.DataFrame({"Risk": sorted(sim_result['risks'])[:100]}))

def generate_multimodal_routes_domestic(route_dist):
    routes = {
        "FULL_AIR": {
            "name": "Full Air",
            "modes": [{"type": "AIR", "percentage": 100}],
            "description": "항공 직배",
            "co2": 1.8 * (route_dist / 1000),
            "icon": ""
        },
        "FULL_ROAD": {
            "name": "Full Road",
            "modes": [{"type": "ROAD", "percentage": 100}],
            "description": "도로 직배",
            "co2": 1.0 * (route_dist / 1000),
            "icon": ""
        },
        "RAIL_DOMINANT": {
            "name": "Rail + Road",
            "modes": [{"type": "RAIL", "percentage": 70}, {"type": "ROAD", "percentage": 30}],
            "description": "철도 + 도로 복합",
            "co2": (0.5 * 0.7 + 1.0 * 0.3) * (route_dist / 1000),
            "icon": ""
        },
        "AIR_ROAD": {
            "name": "Air + Road",
            "modes": [{"type": "AIR", "percentage": 50}, {"type": "ROAD", "percentage": 50}],
            "description": "항공 + 도로 복합",
            "co2": (1.8 * 0.5 + 1.0 * 0.5) * (route_dist / 1000),
            "icon": ""
        },
    }
    return routes

def generate_multimodal_routes_international(route_dist):
    routes = {
        "FULL_SEA": {
            "name": "Full Sea Route",
            "modes": [{"type": "SEA", "percentage": 100}],
            "description": "Port-to-Port direct sea freight",
            "co2": 0.2 * (route_dist / 1000),
            "icon": ""
        },
        "AIR_SEA": {
            "name": "Air + Sea (Fast)",
            "modes": [{"type": "AIR", "percentage": 30}, {"type": "SEA", "percentage": 70}],
            "description": "Air freight to hub + sea transport to destination",
            "co2": (1.8 * 0.3 + 0.2 * 0.7) * (route_dist / 1000),
            "icon": ""
        },
        "SEA_RAIL": {
            "name": "Sea + Rail (Economic)",
            "modes": [{"type": "SEA", "percentage": 60}, {"type": "RAIL", "percentage": 40}],
            "description": "Sea to inland port + rail to destination",
            "co2": (0.2 * 0.6 + 0.5 * 0.4) * (route_dist / 1000),
            "icon": ""
        },
        "SEA_ROAD": {
            "name": "Sea + Road (Flexible)",
            "modes": [{"type": "SEA", "percentage": 75}, {"type": "ROAD", "percentage": 25}],
            "description": "Sea to port + road to final destination",
            "co2": (0.2 * 0.75 + 1.0 * 0.25) * (route_dist / 1000),
            "icon": ""
        },
    }
    return routes

STRATEGY_TEXTS = {
    "ko": {
        "domestic": {
            "STRATEGY_A": {"name": "전략 A: 빠른 배송 (Fast-Track)", "description": "항공 우선 긴급 배송",
                           "pros": ["최단 납기 (시간 단위)", "실시간 추적"], "cons": ["매우 높은 비용"], "suitable_for": "긴급 배송"},
            "STRATEGY_B": {"name": "전략 B: 균형형 (Balanced)", "description": "도로 + 철도 복합 최적화",
                           "pros": ["최적 비용", "정시율 우수"], "cons": ["평균 시간"], "suitable_for": "일반 배송"},
            "STRATEGY_C": {"name": "전략 C: 경제형 (Economic)", "description": "철도 중심 저비용 배송",
                           "pros": ["최저 비용", "탄소 저감"], "cons": ["긴 납기"], "suitable_for": "대량 배송"},
        },
        "international": {
            "STRATEGY_A": {"name": "전략 A: 빠른 배송 (Fast-Track)", "description": "항공 혼합 운송으로 최단 납기",
                           "pros": ["최단 납기", "우선 처리", "Real-time 추적"], "cons": ["높은 비용 (+35%)", "연료비 상승"], "suitable_for": "반도체, 배터리"},
            "STRATEGY_B": {"name": "전략 B: 균형형 (Balanced)", "description": "비용-시간 최적화 표준 운송",
                           "pros": ["최적 비용", "정시율 우수"], "cons": ["평균 리스크"], "suitable_for": "완성차 부품"},
            "STRATEGY_C": {"name": "전략 C: 경제형 (Economic)", "description": "최저 비용 중심 운송",
                           "pros": ["최저 비용 (-22%)", "탄소 저감"], "cons": ["긴 납기"], "suitable_for": "부자재"},
        },
    },
    "en": {
        "domestic": {
            "STRATEGY_A": {"name": "Strategy A: Fast-Track", "description": "Air-priority urgent delivery",
                           "pros": ["Shortest lead time (hours)", "Real-time tracking"], "cons": ["Very high cost"], "suitable_for": "Urgent delivery"},
            "STRATEGY_B": {"name": "Strategy B: Balanced", "description": "Optimized Road + Rail combination",
                           "pros": ["Optimal cost", "Excellent on-time rate"], "cons": ["Average lead time"], "suitable_for": "General delivery"},
            "STRATEGY_C": {"name": "Strategy C: Economic", "description": "Rail-focused low-cost delivery",
                           "pros": ["Lowest cost", "Reduced carbon emissions"], "cons": ["Long lead time"], "suitable_for": "Bulk delivery"},
        },
        "international": {
            "STRATEGY_A": {"name": "Strategy A: Fast-Track", "description": "Air-mixed transport for shortest lead time",
                           "pros": ["Shortest lead time", "Priority handling", "Real-time tracking"], "cons": ["High cost (+35%)", "Higher fuel cost"], "suitable_for": "Semiconductors, batteries"},
            "STRATEGY_B": {"name": "Strategy B: Balanced", "description": "Standard transport optimized for cost & time",
                           "pros": ["Optimal cost", "Excellent on-time rate"], "cons": ["Average risk"], "suitable_for": "Finished vehicle parts"},
            "STRATEGY_C": {"name": "Strategy C: Economic", "description": "Lowest-cost focused transport",
                           "pros": ["Lowest cost (-22%)", "Reduced carbon emissions"], "cons": ["Long lead time"], "suitable_for": "Sub-materials"},
        },
    },
}

def _strategy_text(route_type, strat_key, field):
    lang = st.session_state.get("nav_language", "ko") if hasattr(st, "session_state") else "ko"
    return STRATEGY_TEXTS.get(lang, STRATEGY_TEXTS["ko"])[route_type][strat_key][field]

def generate_default_strategies_domestic(origin, destination, route_dist, cargo_type, dangerous_goods, vehicle_speed):
    rt = "domestic"
    return {
        "STRATEGY_A": {
            "name": _strategy_text(rt, "STRATEGY_A", "name"),
            "description": _strategy_text(rt, "STRATEGY_A", "description"),
            "modes": ["AIR"],
            "time_days": max(1, int(route_dist / (vehicle_speed * 12))),
            "cost_multiplier": 1.5,
            "risk_score": 30,
            "pros": _strategy_text(rt, "STRATEGY_A", "pros"),
            "cons": _strategy_text(rt, "STRATEGY_A", "cons"),
            "suitable_for": _strategy_text(rt, "STRATEGY_A", "suitable_for")
        },
        "STRATEGY_B": {
            "name": _strategy_text(rt, "STRATEGY_B", "name"),
            "description": _strategy_text(rt, "STRATEGY_B", "description"),
            "modes": ["ROAD", "RAIL"],
            "time_days": max(2, int(route_dist / (vehicle_speed * 8))),
            "cost_multiplier": 1.0,
            "risk_score": 45,
            "pros": _strategy_text(rt, "STRATEGY_B", "pros"),
            "cons": _strategy_text(rt, "STRATEGY_B", "cons"),
            "suitable_for": _strategy_text(rt, "STRATEGY_B", "suitable_for")
        },
        "STRATEGY_C": {
            "name": _strategy_text(rt, "STRATEGY_C", "name"),
            "description": _strategy_text(rt, "STRATEGY_C", "description"),
            "modes": ["RAIL"],
            "time_days": max(3, int(route_dist / (vehicle_speed * 6))),
            "cost_multiplier": 0.7,
            "risk_score": 55,
            "pros": _strategy_text(rt, "STRATEGY_C", "pros"),
            "cons": _strategy_text(rt, "STRATEGY_C", "cons"),
            "suitable_for": _strategy_text(rt, "STRATEGY_C", "suitable_for")
        }
    }

def generate_default_strategies_international(origin, destination, route_dist, cargo_type, dangerous_goods, vessel_speed):
    rt = "international"
    return {
        "STRATEGY_A": {
            "name": _strategy_text(rt, "STRATEGY_A", "name"),
            "description": _strategy_text(rt, "STRATEGY_A", "description"),
            "modes": ["AIR", "SEA"],
            "time_days": max(3, int(route_dist / (vessel_speed * 1.852 * 24 * 0.6))),
            "cost_multiplier": 1.35,
            "risk_score": 45,
            "pros": _strategy_text(rt, "STRATEGY_A", "pros"),
            "cons": _strategy_text(rt, "STRATEGY_A", "cons"),
            "suitable_for": _strategy_text(rt, "STRATEGY_A", "suitable_for")
        },
        "STRATEGY_B": {
            "name": _strategy_text(rt, "STRATEGY_B", "name"),
            "description": _strategy_text(rt, "STRATEGY_B", "description"),
            "modes": ["SEA", "RAIL"],
            "time_days": max(7, int(route_dist / (vessel_speed * 1.852 * 24))),
            "cost_multiplier": 1.0,
            "risk_score": 62,
            "pros": _strategy_text(rt, "STRATEGY_B", "pros"),
            "cons": _strategy_text(rt, "STRATEGY_B", "cons"),
            "suitable_for": _strategy_text(rt, "STRATEGY_B", "suitable_for")
        },
        "STRATEGY_C": {
            "name": _strategy_text(rt, "STRATEGY_C", "name"),
            "description": _strategy_text(rt, "STRATEGY_C", "description"),
            "modes": ["SEA", "ROAD"],
            "time_days": max(10, int(route_dist / (vessel_speed * 1.852 * 24 * 1.35))),
            "cost_multiplier": 0.78,
            "risk_score": 75,
            "pros": _strategy_text(rt, "STRATEGY_C", "pros"),
            "cons": _strategy_text(rt, "STRATEGY_C", "cons"),
            "suitable_for": _strategy_text(rt, "STRATEGY_C", "suitable_for")
        }
    }

def run_scenario_simulation(strategy, route_distance, cargo_weight, containers_40, containers_20, dangerous_goods, events_severity):
    iters = 1000

    base_cost = 2000 + (cargo_weight * 0.15 if cargo_weight else 0) + \
                (containers_40 * 3500 + containers_20 * 2200)
    base_cost = base_cost * strategy.get("cost_multiplier", 1.0)

    if dangerous_goods:
        base_cost *= 1.25

    risk_multiplier = 1.0 + (events_severity / 1000.0)
    base_cost = base_cost * risk_multiplier

    costs = np.random.normal(loc=base_cost, scale=base_cost*0.12, size=iters)
    costs = np.abs(costs)

    time_base = strategy.get("time_days", 15)
    times = np.random.normal(loc=time_base, scale=time_base*0.18, size=iters)
    times = np.abs(times)

    risk_base = strategy.get("risk_score", 60)
    risks = np.random.normal(loc=risk_base, scale=15, size=iters)
    risks = np.clip(risks, 0, 100)

    return {
        "costs": costs,
        "times": times,
        "risks": risks,
        "cost_mean": float(np.mean(costs)),
        "cost_std": float(np.std(costs)),
        "time_mean": float(np.mean(times)),
        "time_std": float(np.std(times)),
        "risk_mean": float(np.mean(risks)),
        "risk_std": float(np.std(risks))
    }

def find_nearest_korean_port(domestic_terminal_name):
    origin_coord = DOMESTIC_TERMINALS[domestic_terminal_name][:2]
    korean_ports = {name: info for name, info in PORTS_DB.items() if info[2] == "South Korea"}
    best_name, best_dist = None, float("inf")
    for name, info in korean_ports.items():
        d = haversine_km(origin_coord[0], origin_coord[1], info[0], info[1])
        if d < best_dist:
            best_dist = d
            best_name = name
    return best_name, best_dist

def apply_shipment_mode_adjustments(sim_result, extra_days=0.0, cost_multiplier=1.0):
    new_times = sim_result["times"] + extra_days
    new_costs = sim_result["costs"] * cost_multiplier
    return {
        "costs": new_costs,
        "times": new_times,
        "risks": sim_result["risks"],
        "cost_mean": float(np.mean(new_costs)),
        "cost_std": float(np.std(new_costs)),
        "time_mean": float(np.mean(new_times)),
        "time_std": float(np.std(new_times)),
        "risk_mean": sim_result["risk_mean"],
        "risk_std": sim_result["risk_std"],
    }

def fetch_real_ais_vessels(latmin, latmax, lonmin, lonmax, username):
    if not username:
        return None
    try:
        url = "http://data.aishub.net/ws.php"
        params = {
            "username": username, "format": 1, "output": "json", "compress": 0,
            "latmin": latmin, "latmax": latmax, "lonmin": lonmin, "lonmax": lonmax,
        }
        resp = requests.get(url, params=params, timeout=6)
        data = resp.json()
        if not isinstance(data, list) or len(data) < 2 or not isinstance(data[1], list):
            return None
        vessels = []
        for row in data[1]:
            vessels.append({
                "mmsi": row.get("MMSI", 0),
                "name": row.get("NAME") or f"Vessel {row.get('MMSI', '?')}",
                "type": "Container",
                "lat": float(row.get("LATITUDE", 0) or 0),
                "lon": float(row.get("LONGITUDE", 0) or 0),
                "heading": float(row.get("COG", row.get("HEADING", 0)) or 0),
                "speed": float(row.get("SOG", 0) or 0),
                "status": "Underway",
                "flag": "Unknown",
                "length": 0, "beam": 0,
            })
        return vessels if vessels else None
    except Exception:
        return None

def fetch_exchange_rates(base="USD", targets=("KRW", "EUR", "JPY", "CNY")):
    fallback = {"KRW": 1380.0, "EUR": 0.92, "JPY": 156.0, "CNY": 7.25}
    try:
        url = "https://api.frankfurter.app/latest"
        resp = requests.get(url, params={"from": base, "to": ",".join(targets)}, timeout=5)
        data = resp.json()
        rates = data.get("rates", {})
        return rates if rates else fallback
    except Exception:
        return fallback

def get_exchange_rates_cached():
    if st.session_state.exchange_rates is None:
        st.session_state.exchange_rates = fetch_exchange_rates()
    return st.session_state.exchange_rates

def convert_currency(amount_usd, target):
    if target == "USD":
        return amount_usd
    rate = get_exchange_rates_cached().get(target)
    return amount_usd * rate if rate else amount_usd

def format_currency(amount_usd, target):
    symbols = {"USD": "$", "KRW": "₩", "EUR": "€", "JPY": "¥", "CNY": "¥"}
    converted = convert_currency(amount_usd, target)
    symbol = symbols.get(target, "")
    if target in ("KRW", "JPY"):
        return f"{symbol}{converted:,.0f}"
    return f"{symbol}{converted:,.2f}"

def fetch_port_congestion_index(port_name):
    if port_name in st.session_state.port_congestion_cache:
        return st.session_state.port_congestion_cache[port_name]
    h = int(hashlib.md5(port_name.encode()).hexdigest(), 16)
    congestion_score = 30 + (h % 60)
    waiting_days = round(0.5 + (congestion_score / 100) * 4, 1)
    result = {"congestion_score": congestion_score, "waiting_days": waiting_days, "source": "Estimated (demo)"}
    st.session_state.port_congestion_cache[port_name] = result
    return result

def fetch_typhoon_paths(api_key):
    if not api_key:
        return []
    try:
        url = "http://apis.data.go.kr/1360000/TyphoonInfoService/getTyphoonInfo"
        params = {
            "serviceKey": api_key, "pageNo": 1, "numOfRows": 50, "dataType": "JSON",
            "fromTmFc": (date.today() - timedelta(days=7)).strftime("%Y%m%d0000"),
            "toTmFc": date.today().strftime("%Y%m%d2359"),
        }
        resp = requests.get(url, params=params, timeout=6)
        data = resp.json()
        items = data.get("response", {}).get("body", {}).get("items", {}).get("item", [])
        if isinstance(items, dict):
            items = [items]
        typhoons = []
        for it in items:
            try:
                typhoons.append({
                    "name": it.get("typName", "Unknown"),
                    "lat": float(it.get("typLat", 0)),
                    "lon": float(it.get("typLon", 0)),
                    "wind_speed": it.get("typWs", "-"),
                })
            except Exception:
                continue
        return typhoons
    except Exception:
        return []

def render_typhoon_paths_on_map(m, typhoons):
    for ty in typhoons:
        folium.Marker(
            location=[ty["lat"], ty["lon"]],
            popup=f"태풍 {ty['name']}<br>풍속: {ty['wind_speed']}",
            tooltip=f"{ty['name']}",
            icon=folium.Icon(color="darkred", icon="bolt", prefix="fa"),
        ).add_to(m)
        folium.Circle(
            location=[ty["lat"], ty["lon"]],
            radius=200000, color="#8B0000", fill=True, fillColor="#8B0000", fillOpacity=0.15, weight=1,
        ).add_to(m)

def get_map_view_for_leg(leg_selection, ordered_names, coords_lookup, default_center, default_zoom):
    if leg_selection == "전체 경로" or "→" not in str(leg_selection):
        return default_center, default_zoom
    a_name, b_name = [s.strip() for s in leg_selection.split("→")]
    if a_name not in coords_lookup or b_name not in coords_lookup:
        return default_center, default_zoom
    a, b = coords_lookup[a_name], coords_lookup[b_name]
    center = [(a[1] + b[1]) / 2, (a[0] + b[0]) / 2]
    dist_km = haversine_km(a[0], a[1], b[0], b[1])
    if dist_km < 100:
        zoom = 9
    elif dist_km < 500:
        zoom = 7
    elif dist_km < 2000:
        zoom = 5
    elif dist_km < 6000:
        zoom = 3
    else:
        zoom = 2
    return center, zoom

def render_folium_map(m, height, map_key):
    if ST_FOLIUM_AVAILABLE:
        return st_folium(m, height=height, use_container_width=True, key=map_key,
                          returned_objects=["last_object_clicked_popup"])
    st.components.v1.html(m._repr_html_(), height=height)
    return None

def get_gemini_model(api_key):
    if not GEMINI_AVAILABLE or not api_key:
        return None
    try:
        genai.configure(api_key=api_key)
        return genai.GenerativeModel("gemini-1.5-flash")
    except Exception:
        return None

def ask_gemini(prompt, api_key):
    model = get_gemini_model(api_key)
    if model is None:
        return None
    try:
        resp = model.generate_content(prompt)
        return resp.text
    except Exception as e:
        return f"Gemini 호출 오류: {str(e)[:200]}"

def generate_ai_risk_briefing(news_items, api_key):
    if not news_items:
        return "요약할 뉴스가 없습니다."
    joined = "\n".join(f"- {n}" for n in news_items[:5])
    prompt = (
        "다음은 최신 물류/해운 뉴스 헤드라인입니다. 자동차 부품 국제물류 담당자 관점에서 "
        "이번 주 핵심 리스크를 3줄로 요약하고, 특히 주의해야 할 항로나 요인을 짚어주세요.\n\n" + joined
    )
    result = ask_gemini(prompt, api_key)
    return result or "Gemini API 키가 설정되지 않았거나 호출에 실패했습니다. 사이드바 '전역 설정'에서 API 키를 확인해주세요."

def calculate_customs_clearance_days(incoterms, cargo_type, dangerous_goods):
    base = INCOTERMS_CLEARANCE_DAYS.get(incoterms, {"export_days": 2, "import_days": 4, "note": ""})
    extra = CARGO_CLEARANCE_EXTRA_DAYS.get(cargo_type, 0)
    if dangerous_goods:
        extra += 2
    total = base["export_days"] + base["import_days"] + extra
    return {"export_days": base["export_days"], "import_days": base["import_days"],
            "extra_days": extra, "total_days": total, "note": base["note"]}

def calculate_cbam_cost(co2_tons, carbon_price_eur=EU_CARBON_PRICE_EUR_PER_TON):
    return co2_tons * carbon_price_eur

def estimate_insurance_premium(cargo_value_usd, risk_score):
    base_rate = 0.0015
    risk_rate = (risk_score / 100) * 0.0135
    rate = base_rate + risk_rate
    return cargo_value_usd * rate, rate

def render_forwarder_reputation():
    df = pd.DataFrame(FORWARDER_REPUTATION).T.reset_index().rename(columns={
        "index": "Forwarder", "on_time_rate": "정시율(%)", "incident_rate": "사고율(%)", "avg_rating": "평점(5점)"
    })
    st.dataframe(df, use_container_width=True, hide_index=True)

def _find_korean_font():
    candidates = [
        "C:\\Windows\\Fonts\\malgun.ttf",
        "C:\\Windows\\Fonts\\malgunbd.ttf",
        "C:\\Windows\\Fonts\\NanumGothic.ttf",
        "/System/Library/Fonts/Supplemental/AppleGothic.ttf",
        "/System/Library/Fonts/AppleSDGothicNeo.ttc",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None

def _pdf_write_line(pdf, text, font_name, style, size, height, kind="multicell"):
    pdf.set_x(pdf.l_margin)
    try:
        pdf.set_font(font_name, style, size)
        if kind == "cell":
            pdf.cell(0, height, text, ln=True)
        else:
            pdf.multi_cell(0, height, text)
        return
    except Exception:
        pass

    pdf.set_x(pdf.l_margin)
    try:
        pdf.set_font("Helvetica", style, size)
        safe_text = str(text).encode("latin-1", "replace").decode("latin-1")
        if kind == "cell":
            pdf.cell(0, height, safe_text, ln=True)
        else:
            pdf.multi_cell(0, height, safe_text)
    except Exception:
        pdf.set_x(pdf.l_margin)
        try:
            pdf.set_font("Helvetica", "", size)
            pdf.cell(0, height, "[rendering error - line skipped]", ln=True)
        except Exception:
            pass

def generate_pdf_report(title, route_info: dict, strategy: dict, sim_result: dict):
    if not FPDF_AVAILABLE:
        return None

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    font_path = _find_korean_font()
    font_name = "Helvetica"
    if font_path:
        try:
            pdf.add_font("Korean", "", font_path)
            pdf.add_font("Korean", "B", font_path)
            font_name = "Korean"
        except Exception:
            font_name = "Helvetica"

    def line(text, style="", size=11, height=7, kind="multicell"):
        _pdf_write_line(pdf, str(text), font_name, style, size, height, kind)

    line(title, style="B", size=16, height=10, kind="cell")
    line(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", size=11, height=6, kind="cell")
    if font_name == "Helvetica":
        line("* 시스템에서 한글 폰트를 찾지 못해 일부 문자가 '?'로 표시될 수 있습니다. "
             "Malgun Gothic/NanumGothic 등을 설치하면 한글이 정상 출력됩니다.", size=9, height=5)
    pdf.ln(4)

    line("Route Information", style="B", size=13, height=8, kind="cell")
    for k, v in route_info.items():
        line(f"{k}: {v}", size=11, height=7)
    pdf.ln(2)

    line("Selected Strategy", style="B", size=13, height=8, kind="cell")
    for k, v in strategy.items():
        if isinstance(v, list):
            v = ", ".join(v)
        line(f"{k}: {v}", size=11, height=7)

    if sim_result:
        pdf.ln(2)
        line("Simulation Result (Monte Carlo, 1000 iters)", style="B", size=13, height=8, kind="cell")
        line(f"Expected Cost: ${sim_result['cost_mean']:,.0f} (+/- {sim_result['cost_std']:,.0f})", size=11, height=7, kind="cell")
        line(f"Expected Delivery: {sim_result['time_mean']:.1f} days (+/- {sim_result['time_std']:.1f})", size=11, height=7, kind="cell")
        line(f"Risk Score: {sim_result['risk_mean']:.0f}/100 (+/- {sim_result['risk_std']:.0f})", size=11, height=7, kind="cell")

    try:
        return bytes(pdf.output(dest="S"))
    except TypeError:
        return pdf.output(dest="S").encode("latin-1")

def generate_excel_report(route_info: dict, strategy: dict, sim_result: dict):
    if not OPENPYXL_AVAILABLE and not XLSXWRITER_AVAILABLE:
        return None

    engine = "openpyxl" if OPENPYXL_AVAILABLE else "xlsxwriter"
    buffer = io.BytesIO()
    try:
        with pd.ExcelWriter(buffer, engine=engine) as writer:
            pd.DataFrame(list(route_info.items()), columns=["Item", "Value"]).to_excel(writer, sheet_name="Route", index=False)
            strat_flat = {k: (", ".join(v) if isinstance(v, list) else v) for k, v in strategy.items()}
            pd.DataFrame(list(strat_flat.items()), columns=["Item", "Value"]).to_excel(writer, sheet_name="Strategy", index=False)
            if sim_result:
                sim_df = pd.DataFrame({"Cost": sim_result["costs"], "Time": sim_result["times"], "Risk": sim_result["risks"]})
                sim_df.to_excel(writer, sheet_name="Simulation_Raw", index=False)
        return buffer.getvalue()
    except Exception:
        return None

def build_mailto_link(subject, body, to_addr=""):
    params = urllib.parse.urlencode({"subject": subject, "body": body}, quote_via=urllib.parse.quote)
    return f"mailto:{to_addr}?{params}"

def post_to_slack(webhook_url, message):
    if not webhook_url:
        return False, "Slack Webhook URL이 설정되지 않았습니다 (사이드바 '전역 설정'에서 입력)."
    try:
        resp = requests.post(webhook_url, json={"text": message}, timeout=5)
        if resp.status_code == 200:
            return True, "Slack 전송 성공"
        return False, f"Slack 전송 실패 (status {resp.status_code})"
    except Exception as e:
        return False, f"Slack 전송 오류: {str(e)[:100]}"

def render_report_and_share_section(title, route_info: dict, strategy: dict, sim_result: dict, key_prefix=""):
    with st.expander(t("report_title")):
        col_a, col_b = st.columns(2)

        with col_a:
            if FPDF_AVAILABLE:
                pdf_bytes = generate_pdf_report(title, route_info, strategy, sim_result)
                st.download_button(t("report_pdf_btn"), data=pdf_bytes, file_name="atlas_ai_report.pdf",
                                    mime="application/pdf", key=f"{key_prefix}_pdf_dl")
            else:
                st.caption(t("report_pdf_missing"))

        with col_b:
            excel_bytes = generate_excel_report(route_info, strategy, sim_result)
            if excel_bytes:
                st.download_button(t("report_excel_btn"), data=excel_bytes, file_name="atlas_ai_report.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"{key_prefix}_xlsx_dl")
            else:
                st.caption(t("report_excel_missing"))

        st.divider()
        st.write(t("report_email_header"))
        to_addr = st.text_input(t("report_email_to"), key=f"{key_prefix}_mail_to")
        mail_body = (
            f"{title}\n\n" + "\n".join(f"{k}: {v}" for k, v in route_info.items()) +
            "\n\n[Strategy]\n" + "\n".join(f"{k}: {', '.join(v) if isinstance(v, list) else v}" for k, v in strategy.items())
        )
        mailto = build_mailto_link(title, mail_body, to_addr)
        st.markdown(f'<a href="{mailto}" target="_blank">{t("report_email_open_btn")}</a>', unsafe_allow_html=True)

        st.divider()
        st.write(t("report_slack_header"))
        slack_msg = st.text_area(t("report_slack_msg_label"), value=f"[{title}]\n" + mail_body[:500], key=f"{key_prefix}_slack_msg", height=100)
        if st.button(t("report_slack_send_btn"), key=f"{key_prefix}_slack_btn"):
            ok, msg = post_to_slack(st.session_state.slack_webhook, slack_msg)
            (st.success if ok else st.warning)(msg)

def apply_theme_css():
    if st.session_state.dark_mode:
        st.markdown("""
        <style>
            [data-testid="stAppViewContainer"] { background-color: #0e1117; color: #fafafa; }
            [data-testid="stSidebar"] { background-color: #161a23; }
        </style>
        """, unsafe_allow_html=True)

def cols_adaptive(n):
    if st.session_state.compact_mode:
        return [st.container() for _ in range(n)]
    return st.columns(n)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "atlas_ai_data.db") if "__file__" in globals() else "atlas_ai_data.db"

def get_db_connection():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS favorites (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, type TEXT, origin TEXT,
        destination TEXT, transit TEXT, distance_km REAL, saved_at TEXT)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS simulation_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT, route_type TEXT, origin TEXT,
        destination TEXT, strategy_name TEXT, cost_mean REAL, time_mean REAL, risk_mean REAL)""")
    cur.execute("""CREATE TABLE IF NOT EXISTS active_shipments (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT, route_type TEXT, origin TEXT,
        destination TEXT, strategy_name TEXT, cost_mean REAL, time_mean REAL, risk_mean REAL,
        status TEXT, created_at TEXT)""")
    conn.commit()
    conn.close()

init_db()

def db_save_favorite(name, route_dict):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO favorites (name, type, origin, destination, transit, distance_km, saved_at) VALUES (?,?,?,?,?,?,?)",
        (name, route_dict.get("type", ""), route_dict.get("origin", ""), route_dict.get("destination", ""),
         route_dict.get("transit", ""), route_dict.get("distance_km", 0), datetime.now().strftime("%Y-%m-%d %H:%M"))
    )
    conn.commit()
    conn.close()

def db_load_favorites():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM favorites ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def db_delete_favorite(fav_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM favorites WHERE id=?", (fav_id,))
    conn.commit()
    conn.close()

def db_save_simulation_history(route_type, origin, destination, strategy_name, sim_result):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO simulation_history (timestamp, route_type, origin, destination, strategy_name, cost_mean, time_mean, risk_mean) VALUES (?,?,?,?,?,?,?,?)",
        (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), route_type, origin, destination, strategy_name,
         sim_result["cost_mean"], sim_result["time_mean"], sim_result["risk_mean"])
    )
    conn.commit()
    conn.close()

def db_load_simulation_history(limit=200):
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM simulation_history ORDER BY id DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

def db_save_active_shipment(name, route_type, origin, destination, strategy_name, sim_result):
    conn = get_db_connection()
    conn.execute(
        "INSERT INTO active_shipments (name, route_type, origin, destination, strategy_name, cost_mean, time_mean, risk_mean, status, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (name, route_type, origin, destination, strategy_name, sim_result["cost_mean"], sim_result["time_mean"],
         sim_result["risk_mean"], "대기", datetime.now().strftime("%Y-%m-%d %H:%M"))
    )
    conn.commit()
    conn.close()

def db_load_active_shipments():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM active_shipments ORDER BY id DESC").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def db_update_shipment_status(shipment_id, status):
    conn = get_db_connection()
    conn.execute("UPDATE active_shipments SET status=? WHERE id=?", (status, shipment_id))
    conn.commit()
    conn.close()

def db_delete_shipment(shipment_id):
    conn = get_db_connection()
    conn.execute("DELETE FROM active_shipments WHERE id=?", (shipment_id,))
    conn.commit()
    conn.close()

def render_favorites_manager(current_route: dict, key_prefix=""):
    with st.expander(t("fav_title")):
        default_name = f"{current_route.get('origin', '')} → {current_route.get('destination', '')}"
        fav_name = st.text_input(t("fav_name_input"), value=default_name, key=f"{key_prefix}_fav_name")
        if st.button(t("fav_save_btn"), key=f"{key_prefix}_fav_save"):
            db_save_favorite(fav_name, current_route)
            st.success(f"'{fav_name}' {t('fav_saved_msg')}")
            st.rerun()

        favorites = db_load_favorites()
        if favorites:
            fav_df = pd.DataFrame(favorites)[["id", "name", "type", "origin", "destination", "distance_km", "saved_at"]]
            st.dataframe(fav_df, use_container_width=True, hide_index=True)
            del_id = st.selectbox(
                t("fav_delete_select"), options=[f["id"] for f in favorites],
                format_func=lambda i: next(f["name"] for f in favorites if f["id"] == i), key=f"{key_prefix}_fav_del_sel"
            )
            if st.button(t("fav_delete_btn"), key=f"{key_prefix}_fav_del_btn"):
                db_delete_favorite(del_id)
                st.rerun()
        else:
            st.caption(t("fav_empty"))

HS_CODE_TARIFF_TABLE = {
    "8708": {"desc": "자동차 부품", "rate": 8.0},
    "8703": {"desc": "승용차", "rate": 8.0},
    "8507": {"desc": "축전지(배터리)", "rate": 8.0},
    "8541": {"desc": "반도체 소자", "rate": 0.0},
    "8542": {"desc": "집적회로", "rate": 0.0},
    "4011": {"desc": "타이어", "rate": 5.0},
    "7007": {"desc": "안전유리", "rate": 8.0},
    "8544": {"desc": "전선/케이블", "rate": 6.5},
}

def fetch_customs_tariff_rate(hs_code, api_key):
    hs4 = (hs_code or "")[:4]
    if api_key:
        try:
            url = "https://unipass.customs.go.kr/openapi/service/trifFxrtInfoQry/getTrifFxrtInfo"
            resp = requests.get(url, params={"crkyCn": api_key, "hsSgn": hs_code}, timeout=6)
            if resp.status_code == 200 and resp.text.strip():
                return {"rate": None, "raw_response": resp.text[:300], "source": "관세청 unipass API"}
        except Exception:
            pass
    fallback = HS_CODE_TARIFF_TABLE.get(hs4, {"desc": "일반품목(미등록 HS코드, 근사치)", "rate": 8.0})
    return {"rate": fallback["rate"], "desc": fallback["desc"], "source": "Estimated (demo table)"}

def calculate_tco(freight_cost, customs_duty, cbam_cost, insurance_premium):
    return freight_cost + customs_duty + cbam_cost + insurance_premium

def render_strategy_comparison(strategies, route_distance, cargo_weight, container_40, container_20, dangerous_goods, key_prefix=""):
    with st.expander(t("compare_strategy_title")):
        if st.button(t("compare_strategy_btn"), key=f"{key_prefix}_compare_btn"):
            rows = []
            for strat_key, strat in strategies.items():
                res = run_scenario_simulation(strat, route_distance, cargo_weight, container_40, container_20, dangerous_goods, 0)
                rows.append({
                    "Strategy": strat.get("name", strat_key),
                    "Cost (USD)": round(res["cost_mean"], 0),
                    "Time (days)": round(res["time_mean"], 1),
                    "Risk (0-100)": round(res["risk_mean"], 0),
                })
            df = pd.DataFrame(rows)
            st.dataframe(df, use_container_width=True, hide_index=True)

            melted = df.melt(id_vars="Strategy", var_name="Metric", value_name="Value")
            chart = (
                alt.Chart(melted)
                .mark_bar()
                .encode(
                    x=alt.X("Strategy:N", axis=alt.Axis(labelAngle=0, title=None)),
                    y=alt.Y("Value:Q"),
                    color="Strategy:N",
                    column=alt.Column("Metric:N", header=alt.Header(labelAngle=0)),
                    tooltip=["Strategy", "Metric", "Value"],
                )
                .properties(width=150, height=250)
            )
            st.altair_chart(chart, use_container_width=False)

def render_route_comparison(coords_lookup, origin_name, key_prefix, is_domestic):
    with st.expander(t("compare_route_title")):
        other_options = [n for n in coords_lookup.keys() if n != origin_name]
        fmt = display_terminal_name if is_domestic else (lambda x: x)
        compare_dest = st.selectbox(t("compare_route_dest_label"), other_options, format_func=fmt, key=f"{key_prefix}_compare_dest")
        if st.button(t("compare_route_btn"), key=f"{key_prefix}_compare_route_btn"):
            oc = coords_lookup[origin_name]
            dc = coords_lookup[compare_dest]
            dist = haversine_km(oc[0], oc[1], dc[0], dc[1])
            if is_domestic:
                strat = generate_default_strategies_domestic(origin_name, compare_dest, dist, "일반화물", False, 80)["STRATEGY_B"]
            else:
                strat = generate_default_strategies_international(origin_name, compare_dest, dist, "General Cargo", False, 12)["STRATEGY_B"]
            res = run_scenario_simulation(strat, dist, 10.0, 1, 0, False, 0)
            comp_df = pd.DataFrame([
                {"Destination": compare_dest, "Distance(km)": round(dist, 0), "Est.Cost(USD)": round(res["cost_mean"], 0),
                 "Est.Time(days)": round(res["time_mean"], 1), "Risk": round(res["risk_mean"], 0)},
            ])
            st.dataframe(comp_df, use_container_width=True, hide_index=True)
            st.caption(f"* '{origin_name} → {compare_dest}' — estimated using the Balanced (B) strategy.")

def suggest_alternative_strategy(strategies, current_key):
    current_risk = strategies.get(current_key, {}).get("risk_score", 100)
    candidates = [(k, v) for k, v in strategies.items() if k != current_key and v.get("risk_score", 100) < current_risk]
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[1]["risk_score"])
    return candidates[0]

def check_and_alert_risk(sim_result, strategies, current_strategy_key, route_label, key_prefix=""):
    threshold = st.session_state.risk_alert_threshold
    if sim_result["risk_mean"] <= threshold:
        return
    st.error(f"리스크 스코어({sim_result['risk_mean']:.0f})가 설정된 임계값({threshold})을 초과했습니다!")

    alt = suggest_alternative_strategy(strategies, current_strategy_key)
    if alt:
        alt_key, alt_strat = alt
        st.info(f"대안 제안: **{alt_strat.get('name', alt_key)}** (리스크 {alt_strat.get('risk_score', '-')}/100, "
                f"예상 {alt_strat.get('time_days', '-')}일, 비용 {alt_strat.get('cost_multiplier', 1.0):.2f}x)")

    if st.session_state.auto_slack_alert and st.session_state.slack_webhook:
        msg = f"[ATLAS AI] {route_label} 리스크 경고: {sim_result['risk_mean']:.0f}/100 (임계값 {threshold})"
        ok, _ = post_to_slack(st.session_state.slack_webhook, msg)
        if ok:
            st.caption("Slack으로 자동 알림을 전송했습니다.")

def evaluate_shipment_decision(sim_result, quoted_price):
    if quoted_price <= 0:
        return None
    costs = sim_result["costs"]
    profit_prob = float((quoted_price > costs).mean()) * 100
    margin = quoted_price - sim_result["cost_mean"]
    margin_rate = (margin / quoted_price * 100) if quoted_price else 0
    if profit_prob >= 70 and margin_rate >= 10:
        recommendation, verdict = "수락 추천", "accept"
    elif profit_prob >= 50:
        recommendation, verdict = "조건부 수락 (모니터링 필요)", "caution"
    else:
        recommendation, verdict = "비추천 (손실 가능성 높음)", "reject"
    return {"profit_prob": profit_prob, "margin": margin, "margin_rate": margin_rate,
            "recommendation": recommendation, "verdict": verdict}

def render_accept_reject_section(sim_result, key_prefix=""):
    with st.expander(t("accept_reject_title")):
        if not sim_result:
            st.caption(t("accept_reject_no_sim"))
            return
        default_price = float(round(sim_result["cost_mean"] * 1.15, -2)) if sim_result["cost_mean"] else 0.0
        quoted_price = st.number_input(t("accept_reject_price_label"), min_value=0.0, value=default_price, step=100.0, key=f"{key_prefix}_quoted_price")
        if quoted_price > 0:
            decision = evaluate_shipment_decision(sim_result, quoted_price)
            c1, c2, c3 = cols_adaptive(3)
            with c1:
                st.metric(t("accept_reject_margin_label"), format_currency(decision["margin"], st.session_state.currency), delta=f"{decision['margin_rate']:.1f}%")
            with c2:
                st.metric(t("accept_reject_prob_label"), f"{decision['profit_prob']:.0f}%")
            with c3:
                st.metric(t("accept_reject_verdict_label"), decision["recommendation"])
            st.caption(t("accept_reject_caption"))

def recommend_repositioning(current_terminal, max_km=150):
    if current_terminal not in DOMESTIC_TERMINALS:
        return None
    cur_coord = DOMESTIC_TERMINALS[current_terminal][:2]
    candidates = []
    for name, info in DOMESTIC_TERMINALS.items():
        if name == current_terminal:
            continue
        dist = haversine_km(cur_coord[0], cur_coord[1], info[0], info[1])
        if dist > max_km:
            continue
        cg = fetch_port_congestion_index(name)
        candidates.append((name, dist, cg["congestion_score"]))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[2])
    return candidates[0]

def render_repositioning_section(destination_input):
    with st.expander(t("reposition_title")):
        rec = recommend_repositioning(destination_input)
        if rec:
            name, dist, score = rec
            st.info(t("reposition_result_msg").format(
                dest=display_terminal_name(destination_input), name=display_terminal_name(name), dist=dist, score=score))
        else:
            st.caption(t("reposition_empty"))

def calculate_social_benefit_road_vs_rail(route_dist):
    road_co2 = TRANSPORT_MODES["ROAD"]["co2"] * (route_dist / 1000)
    rail_co2 = TRANSPORT_MODES["RAIL"]["co2"] * (route_dist / 1000)
    co2_reduction = road_co2 - rail_co2
    air_pollution_reduction_pct = (co2_reduction / road_co2 * 100) if road_co2 else 0
    congestion_relief_km = route_dist * 0.02
    return {
        "road_co2": road_co2, "rail_co2": rail_co2, "co2_reduction": co2_reduction,
        "air_pollution_reduction_pct": air_pollution_reduction_pct,
        "accident_risk_road": TRANSPORT_MODES["ROAD"]["risk"], "accident_risk_rail": TRANSPORT_MODES["RAIL"]["risk"],
        "congestion_relief_km": congestion_relief_km,
    }

def render_social_benefit_section(route_dist):
    with st.expander(t("social_benefit_title")):
        sb = calculate_social_benefit_road_vs_rail(route_dist)
        b1, b2, b3, b4 = cols_adaptive(4)
        with b1:
            st.metric(t("social_benefit_co2_label"), f"{sb['co2_reduction']:.2f}t",
                       delta=t("social_benefit_co2_delta").format(road=sb['road_co2'], rail=sb['rail_co2']))
        with b2:
            st.metric(t("social_benefit_air_label"), f"{sb['air_pollution_reduction_pct']:.0f}%")
        with b3:
            st.metric(t("social_benefit_accident_label"),
                       t("social_benefit_accident_delta").format(road=sb['accident_risk_road'], rail=sb['accident_risk_rail']))
        with b4:
            st.metric(t("social_benefit_congestion_label"), f"{sb['congestion_relief_km']:.0f} {t('social_benefit_congestion_unit')}")
        st.caption(t("social_benefit_caption"))

def fetch_freight_car_availability(terminal_name):
    cache = st.session_state.setdefault("freight_car_cache", {})
    if terminal_name in cache:
        return cache[terminal_name]
    h = int(hashlib.md5((terminal_name + "_car").encode()).hexdigest(), 16)
    result = {"available_cars": 5 + (h % 40), "next_train_min": 15 + (h % 180)}
    cache[terminal_name] = result
    return result

def render_rail_logistics_platform():
    with st.expander(t("rail_platform_title")):
        st.write(t("rail_platform_cars_header"))
        car_rows = []
        for name in list(DOMESTIC_TERMINALS.keys())[:10]:
            info = fetch_freight_car_availability(name)
            car_rows.append({t("rail_col_terminal"): name, t("rail_col_cars"): info["available_cars"], t("rail_col_next"): info["next_train_min"]})
        st.dataframe(pd.DataFrame(car_rows), use_container_width=True, hide_index=True)

        st.divider()
        st.write(t("rail_platform_ktx_header"))
        ktx_origin = st.selectbox(t("rail_platform_ktx_origin"), list(DOMESTIC_TERMINALS.keys()), format_func=display_terminal_name, key="ktx_origin")
        ktx_dest = st.selectbox(t("rail_platform_ktx_dest"), [n for n in DOMESTIC_TERMINALS.keys() if n != ktx_origin], format_func=display_terminal_name, key="ktx_dest")
        ktx_desc = st.text_input(t("rail_platform_ktx_desc"), key="ktx_desc", placeholder=t("ktx_placeholder"))
        if st.button(t("rail_platform_ktx_submit"), key="ktx_submit_btn"):
            o, d = DOMESTIC_TERMINALS[ktx_origin][:2], DOMESTIC_TERMINALS[ktx_dest][:2]
            ktx_dist = haversine_km(o[0], o[1], d[0], d[1])
            ktx_sim = {"cost_mean": max(50000.0, ktx_dist * 800), "time_mean": 0.3, "risk_mean": 15}
            db_save_active_shipment(f"KTX특송 {ktx_origin}→{ktx_dest}", "철도특송", ktx_origin, ktx_dest, "KTX 특송 (당일)", ktx_sim)
            st.success(t("ktx_success_msg").format(desc=ktx_desc or t("ktx_no_desc")))

def render_pattern_analysis():
    with st.expander(t("dashboard_pattern_title")):
        history = db_load_simulation_history(500)
        if not history:
            st.caption(t("dashboard_pattern_empty"))
            return
        df = pd.DataFrame(history)
        count_col = t("label_count")
        top_routes = (df.groupby(["origin", "destination"]).size()
                      .reset_index(name=count_col).sort_values(count_col, ascending=False).head(5))
        top_strategy = df["strategy_name"].value_counts().head(5)
        avg_risk_by_type = df.groupby("route_type")["risk_mean"].mean()

        st.write(t("dashboard_pattern_top_routes"))
        st.dataframe(top_routes, use_container_width=True, hide_index=True)

        col_a, col_b = st.columns(2)
        with col_a:
            st.write(t("dashboard_pattern_top_strategy"))
            bar_chart_horizontal_labels(top_strategy, t("label_strategy"), t("label_count"))
        with col_b:
            st.write(t("dashboard_pattern_avg_risk"))
            bar_chart_horizontal_labels(avg_risk_by_type, t("label_route_type"), t("label_avg_risk"))

        if GEMINI_AVAILABLE and st.session_state.gemini_api_key:
            if st.button(t("dashboard_pattern_ai_btn"), key="pattern_ai_btn"):
                summary_text = (f"Top 5 routes: {top_routes.to_dict('records')}\n"
                                 f"Preferred strategies: {top_strategy.to_dict()}\n"
                                 f"Avg risk by type: {avg_risk_by_type.to_dict()}")
                prompt = ("The following is a summary of logistics operation history data. "
                          "Please summarize the tacit knowledge (operator experience/know-how) visible in this pattern in 3 points.\n\n" + summary_text)
                with st.spinner(t("pattern_ai_spinner")):
                    insight = ask_gemini(prompt, st.session_state.gemini_api_key)
                st.info(insight or t("pattern_ai_no_response"))
        else:
            st.caption(t("pattern_ai_hint"))

def suggest_delay_buffer(sim_result):
    if not sim_result or not sim_result.get("time_mean"):
        return None
    variability = sim_result["time_std"] / sim_result["time_mean"]
    if variability > 0.2:
        return round(sim_result["time_std"] * 1.5, 1)
    return None

def render_delay_buffer_warning(sim_result):
    buf = suggest_delay_buffer(sim_result)
    if buf:
        st.warning(t("delay_buffer_warning").format(std=sim_result['time_std'], buf=buf))

def render_supply_chain_visibility():
    shipments = db_load_active_shipments()
    if not shipments:
        return
    with st.expander(t("dashboard_visibility_title")):
        dest_counts = pd.Series([s["destination"] for s in shipments]).value_counts()
        bar_chart_horizontal_labels(dest_counts, t("label_destination"), t("label_count"))
        type_counts = pd.Series([s["route_type"] for s in shipments]).value_counts()
        bar_chart_horizontal_labels(type_counts, t("label_route_type"), t("label_count"))

def render_dashboard_tab():
    st.header(t("dashboard_header"))

    shipments = db_load_active_shipments()
    history = db_load_simulation_history(200)

    total_count = len(shipments)
    avg_risk = sum(s["risk_mean"] for s in shipments) / total_count if total_count else 0
    total_cost = sum(s["cost_mean"] for s in shipments)

    kc1, kc2, kc3, kc4 = cols_adaptive(4)
    with kc1:
        st.metric(t("dashboard_kpi_shipments"), total_count)
    with kc2:
        st.metric(t("dashboard_kpi_avg_risk"), f"{avg_risk:.0f}/100")
    with kc3:
        st.metric(t("dashboard_kpi_total_cost"), format_currency(total_cost, st.session_state.currency))
    with kc4:
        st.metric(t("dashboard_kpi_sim_count"), len(history))

    st.divider()

    st.subheader(t("dashboard_shipments_title"))
    if shipments:
        ship_df = pd.DataFrame(shipments)
        st.dataframe(
            ship_df[["id", "name", "route_type", "origin", "destination", "strategy_name", "cost_mean", "time_mean", "risk_mean", "status", "created_at"]],
            use_container_width=True, hide_index=True
        )
        STATUS_OPTIONS = ["대기", "진행중", "도착완료", "지연"]
        STATUS_LABELS_EN = {"대기": "Pending", "진행중": "In Progress", "도착완료": "Arrived", "지연": "Delayed"}
        status_fmt = lambda s: s if st.session_state.nav_language == "ko" else STATUS_LABELS_EN.get(s, s)

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            target_id = st.selectbox(t("dashboard_status_change_label"), options=[s["id"] for s in shipments],
                                      format_func=lambda i: next(s["name"] for s in shipments if s["id"] == i), key="dash_status_sel")
        with col_b:
            new_status = st.selectbox(t("dashboard_status_new_label"), STATUS_OPTIONS, format_func=status_fmt, key="dash_status_val")
        with col_c:
            st.write("")
            st.write("")
            if st.button(t("dashboard_status_update_btn"), key="dash_status_btn"):
                db_update_shipment_status(target_id, new_status)
                st.rerun()

        del_id = st.selectbox(t("dashboard_delete_label"), options=[s["id"] for s in shipments],
                               format_func=lambda i: next(s["name"] for s in shipments if s["id"] == i), key="dash_del_sel")
        if st.button(t("dashboard_delete_btn"), key="dash_del_btn"):
            db_delete_shipment(del_id)
            st.rerun()
    else:
        st.caption(t("dashboard_shipments_empty"))

    st.divider()

    st.subheader(t("dashboard_cost_trend_title"))
    if history:
        hist_df = pd.DataFrame(history).sort_values("id")
        chart = alt.Chart(hist_df).mark_line(point=True).encode(
            x=alt.X("timestamp:N", axis=alt.Axis(labelAngle=0, title=None)),
            y=alt.Y("cost_mean:Q", title="Expected Cost (USD)"),
            color="route_type:N",
            tooltip=["timestamp", "route_type", "origin", "destination", "strategy_name", "cost_mean"],
        ).properties(height=300)
        st.altair_chart(chart, use_container_width=True)
    else:
        st.caption(t("dashboard_cost_trend_empty"))

    st.divider()
    render_supply_chain_visibility()
    render_pattern_analysis()

def render_global_settings_sidebar():
    with st.sidebar:
        lang_display = {"ko": "한국어", "en": "English"}
        st.session_state.nav_language = st.radio(
            "Language / 언어", options=["ko", "en"], format_func=lambda x: lang_display[x],
            index=["ko", "en"].index(st.session_state.nav_language), horizontal=True, key="cfg_lang"
        )

        with st.expander(t("sidebar_global_settings"), expanded=False):
            st.caption("API 키를 입력하면 실데이터로 자동 전환됩니다. 비워두면 시뮬레이션 데이터가 사용돼요.")
            st.session_state.aishub_username = st.text_input(
                "AISHub Username", value=st.session_state.aishub_username, key="cfg_aishub")
            st.session_state.kma_api_key = st.text_input(
                "기상청(data.go.kr) API Key", value=st.session_state.kma_api_key, type="password", key="cfg_kma")
            st.session_state.gemini_api_key = st.text_input(
                "Gemini API Key", value=st.session_state.gemini_api_key, type="password", key="cfg_gemini")
            st.session_state.slack_webhook = st.text_input(
                "Slack Webhook URL", value=st.session_state.slack_webhook, key="cfg_slack")
            st.session_state.customs_api_key = st.text_input(
                "관세청(unipass) API Key", value=st.session_state.customs_api_key, type="password", key="cfg_customs")

            st.divider()
            currency_options = ["USD", "KRW", "EUR", "JPY", "CNY"]
            st.session_state.currency = st.selectbox(
                "표시 통화", currency_options, index=currency_options.index(st.session_state.currency), key="cfg_currency")
            st.session_state.dark_mode = st.toggle("다크 모드", value=st.session_state.dark_mode, key="cfg_dark")
            st.session_state.compact_mode = st.toggle("컴팩트(모바일) 모드", value=st.session_state.compact_mode, key="cfg_compact")

            st.divider()
            st.write("**리스크 자동 알림**")
            st.session_state.risk_alert_threshold = st.slider(
                "리스크 스코어 경고 임계값", min_value=0, max_value=100, value=st.session_state.risk_alert_threshold, key="cfg_risk_threshold")
            st.session_state.auto_slack_alert = st.toggle(
                "임계값 초과 시 자동 Slack 알림", value=st.session_state.auto_slack_alert, key="cfg_auto_slack")

            if not ST_FOLIUM_AVAILABLE:
                st.caption("`pip install streamlit-folium` 설치 시 지도 클릭 인터랙션이 추가됩니다.")
            if not FPDF_AVAILABLE:
                st.caption("`pip install fpdf2` 설치 시 PDF 리포트 다운로드가 활성화됩니다.")
            if not (OPENPYXL_AVAILABLE or XLSXWRITER_AVAILABLE):
                st.caption("`pip install openpyxl` 설치 시 Excel 리포트 다운로드가 활성화됩니다.")

apply_theme_css()
render_global_settings_sidebar()

tab0, tab1, tab2 = st.tabs([t("tab_dashboard"), t("tab_domestic"), t("tab_intl")])

with tab0:
    render_dashboard_tab()

with tab1:
    st.header(t("domestic_header"))

    render_ai_dashboard()

    st.divider()

    with st.sidebar.expander(t("sidebar_domestic_settings"), expanded=True):
        _domestic_cargo_options = ["자동차부품", "전자제품", "식품", "일반화물", "위험물"]
        _domestic_cargo_labels_en = {"자동차부품": "Automotive Parts", "전자제품": "Electronics", "식품": "Food",
                                      "일반화물": "General Cargo", "위험물": "Dangerous Goods"}
        cargo_type = st.selectbox(
            t("cargo_type_label"), _domestic_cargo_options,
            format_func=lambda x: x if st.session_state.nav_language == "ko" else _domestic_cargo_labels_en.get(x, x),
            key="domestic_cargo"
        )

        st.divider()
        st.write("**Cargo Details**")

        is_bulk_cargo = st.checkbox("Bulk Cargo", value=False, key="domestic_bulk")

        if is_bulk_cargo:
            st.write("**Bulk Cargo Settings**")
            bulk_type = st.selectbox("Bulk Cargo Type", BULK_CARGO_TYPES, key="domestic_bulk_type")

            col_bulk1, col_bulk2 = st.columns(2)
            with col_bulk1:
                bulk_density = st.number_input("Density (ton/m³)", min_value=0.1, value=1.5, step=0.1, key="domestic_bulk_density")
            with col_bulk2:
                bulk_hazard = st.selectbox("Hazard Level", ["Low", "Medium", "High"], key="domestic_bulk_hazard")

            bulk_stowage = st.selectbox("Stowage", ["Special", "Ventilated", "Certified", "Standard"], key="domestic_bulk_stowage")
            bulk_quantity = st.number_input("Quantity (tons)", min_value=0.0, value=100.0, step=10.0, key="domestic_bulk_qty")
            bulk_volume = bulk_quantity / bulk_density if bulk_density > 0 else 0
            st.write(f"**Estimated Volume: {bulk_volume:.2f} m³**")

            cargo_weight = bulk_quantity
            container_40 = 0
            container_20 = 0
        else:
            commodity = st.text_area("Commodity Description", value="", height=50, placeholder="화물 설명 입력", key="domestic_commodity")
            cargo_weight = st.number_input("Total Weight (tons)", min_value=0.0, value=10.0, step=0.5, key="domestic_weight")
            container_40 = st.number_input('40ft Containers', min_value=0, max_value=100, value=1, key="domestic_40ft")
            container_20 = st.number_input('20ft Containers', min_value=0, max_value=100, value=0, key="domestic_20ft")

        dangerous_goods = st.checkbox(t("dangerous_goods_label"), value=False, key="domestic_danger")

        st.divider()
        st.write("**Schedule**")
        dep_date = st.date_input("Departure Date", value=date.today(), key="domestic_dep_date")
        dep_time = st.time_input("Departure Time", value=dtime(hour=9, minute=0), key="domestic_dep_time")
        desired_arrival = st.date_input("Desired Arrival", value=date.today() + timedelta(days=3), key="domestic_arrival")

        st.divider()
        st.write("**Route Selection**")

        origin_input = st.selectbox("Origin (출발지)", list(DOMESTIC_TERMINALS.keys()),
                                     format_func=display_terminal_name, key="domestic_origin")

        transit_options = [tm for tm in DOMESTIC_TERMINALS.keys() if tm != origin_input]
        transit_inputs = st.multiselect("Transit (경유지 - 선택)", transit_options,
                                         format_func=display_terminal_name, key="domestic_transit")

        destination_input = st.selectbox("Destination (도착지)", [tm for tm in DOMESTIC_TERMINALS.keys() if tm != origin_input],
                                          format_func=display_terminal_name, key="domestic_destination")

        st.divider()
        st.write("**Transportation Modes**")
        use_air = st.checkbox("Air Transport", value=False, key="domestic_air")
        use_road = st.checkbox("Road Transport", value=True, key="domestic_road")
        use_rail = st.checkbox("Rail Transport", value=False, key="domestic_rail")
        use_sea = st.checkbox("Sea Transport", value=False, key="domestic_sea")

        st.divider()
        st.write("**Settings**")
        vehicle_speed = st.slider("Average Speed (km/h)", 40, 100, 80, key="domestic_speed")
        show_map_domestic = st.checkbox("Show Interactive Map", value=True, key="domestic_show_map")

    origin_coords = DOMESTIC_TERMINALS[origin_input][:2]
    destination_coords = DOMESTIC_TERMINALS[destination_input][:2]
    transit_coords = [DOMESTIC_TERMINALS[tm][:2] for tm in transit_inputs] if transit_inputs else []

    total_distance = haversine_km(origin_coords[0], origin_coords[1],
                                  transit_coords[0][0] if transit_coords else destination_coords[0],
                                  transit_coords[0][1] if transit_coords else destination_coords[1])

    for i in range(len(transit_coords) - 1):
        total_distance += haversine_km(transit_coords[i][0], transit_coords[i][1],
                                      transit_coords[i+1][0], transit_coords[i+1][1])

    if transit_coords:
        total_distance += haversine_km(transit_coords[-1][0], transit_coords[-1][1],
                                      destination_coords[0], destination_coords[1])

    ordered_names_domestic = [origin_input] + transit_inputs + [destination_input]
    domestic_coords_lookup = {k: v[:2] for k, v in DOMESTIC_TERMINALS.items()}

    if st.session_state.ais_vessels_domestic is None:
        real_vessels = None
        if st.session_state.aishub_username:
            lats = [domestic_coords_lookup[n][1] for n in ordered_names_domestic]
            lons = [domestic_coords_lookup[n][0] for n in ordered_names_domestic]
            real_vessels = fetch_real_ais_vessels(
                min(lats) - 1, max(lats) + 1, min(lons) - 1, max(lons) + 1, st.session_state.aishub_username
            )
        st.session_state.ais_vessels_domestic = real_vessels or generate_realistic_ais_vessels_domestic(60)
        st.session_state.ais_source_domestic = "AISHub (실데이터)" if real_vessels else "시뮬레이션"

    typhoons_domestic = fetch_typhoon_paths(st.session_state.kma_api_key) if st.session_state.kma_api_key else []

    if show_map_domestic:
        st.subheader("Real-time Domestic Logistics Map")
        vessel_source_label = st.session_state.get("ais_source_domestic", "시뮬레이션")
        st.markdown(f"**Active Vessels: {len(st.session_state.ais_vessels_domestic)}+ ({vessel_source_label}) | Terminals: 22 | Risk Zones | ━━ Routes**")

        leg_options_domestic = ["전체 경로"] + [f"{ordered_names_domestic[i]} → {ordered_names_domestic[i+1]}"
                                                  for i in range(len(ordered_names_domestic) - 1)]
        st.session_state.map_zoom_leg_domestic = st.selectbox(
            "구간 확대", leg_options_domestic, key="domestic_leg_zoom",
            index=leg_options_domestic.index(st.session_state.map_zoom_leg_domestic)
            if st.session_state.map_zoom_leg_domestic in leg_options_domestic else 0
        )

        default_center = [(origin_coords[1] + destination_coords[1]) / 2, (origin_coords[0] + destination_coords[0]) / 2]
        map_center, map_zoom = get_map_view_for_leg(
            st.session_state.map_zoom_leg_domestic, ordered_names_domestic, domestic_coords_lookup, default_center, 7
        )

        m = folium.Map(
            location=map_center,
            zoom_start=map_zoom,
            tiles="CartoDB positron"
        )

        MiniMap().add_to(m)
        Fullscreen().add_to(m)

        route_waypoints = [[DOMESTIC_TERMINALS[tm][1], DOMESTIC_TERMINALS[tm][0]] for tm in transit_inputs]
        base_route = [[origin_coords[1], origin_coords[0]]] + route_waypoints + [[destination_coords[1], destination_coords[0]]]

        folium.PolyLine(
            locations=base_route,
            color="#0077BE",
            weight=3,
            opacity=0.8,
            popup="Base Route",
            tooltip="Base Route",
            dash_array=None
        ).add_to(m)

        ai_route = create_ai_route_with_waypoints(origin_coords,
                                                  [DOMESTIC_TERMINALS[tm][:2] for tm in transit_inputs],
                                                  destination_coords,
                                                  num_points=15, offset_km=20)

        folium.PolyLine(
            locations=ai_route,
            color="#FF7F0E",
            weight=3,
            opacity=0.7,
            popup="AI Optimized Route",
            tooltip="AI Route (Optimized)",
            dash_array="5, 5"
        ).add_to(m)

        for term_name, term_info in DOMESTIC_TERMINALS.items():
            lon, lat, term_type, category = term_info

            if term_name == origin_input:
                color = "#1f77b4"
            elif term_name == destination_input:
                color = "#ff7f0e"
            elif term_name in transit_inputs:
                color = "#2ca02c"
            else:
                color = "#95a5a6"

            folium.CircleMarker(
                location=[lat, lon],
                radius=5,
                color=color,
                fill=True,
                fillColor=color,
                fillOpacity=0.8,
                weight=1.5,
                popup=f"<b>{term_name}</b><br>{term_type}<br>{category}",
                tooltip=term_name
            ).add_to(m)

        render_ais_vessels_on_map(m, st.session_state.ais_vessels_domestic)
        render_typhoon_paths_on_map(m, typhoons_domestic)

        render_folium_map(m, height=750, map_key="domestic_map")

        if typhoons_domestic:
            st.warning(f"활성 태풍 {len(typhoons_domestic)}건 감지 - 해상 경로 리스크 참고")
        elif st.session_state.kma_api_key:
            st.caption("현재 활성 태풍 정보 없음")

        cong_col1, cong_col2 = cols_adaptive(2)
        with cong_col1:
            if DOMESTIC_TERMINALS[origin_input][2] == "Port":
                cg = fetch_port_congestion_index(origin_input)
                st.metric(f"{origin_input} 혼잡도", f"{cg['congestion_score']}/100", delta=f"대기 {cg['waiting_days']}일")
        with cong_col2:
            if DOMESTIC_TERMINALS[destination_input][2] == "Port":
                cg = fetch_port_congestion_index(destination_input)
                st.metric(f"{destination_input} 혼잡도", f"{cg['congestion_score']}/100", delta=f"대기 {cg['waiting_days']}일")

    render_ais_fleet_statistics(st.session_state.ais_vessels_domestic, coverage_label="Domestic")

    st.subheader("Fleet Statistics")
    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)

    with col_stat1:
        st.metric("Total Terminals", len(DOMESTIC_TERMINALS))
    with col_stat2:
        st.metric("Airports", len([tm for tm in DOMESTIC_TERMINALS.values() if tm[2] == "Airport"]))
    with col_stat3:
        st.metric("Ports", len([tm for tm in DOMESTIC_TERMINALS.values() if tm[2] == "Port"]))
    with col_stat4:
        st.metric("Terminals", len([tm for tm in DOMESTIC_TERMINALS.values() if tm[2] == "Terminal"]))

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("총거리", f"{total_distance:.0f} km")
    with col2:
        est_time = max(1, int(total_distance / vehicle_speed * 24))
        st.metric("예상시간", f"{est_time}시간")
    with col3:
        st.metric("화물", f"{cargo_weight}톤")

    st.subheader("Multi-Modal Transportation Routes")
    multimodal_routes = generate_multimodal_routes_domestic(total_distance)

    cols = st.columns(len(multimodal_routes))

    for idx, (route_key, route_info) in enumerate(multimodal_routes.items()):
        with cols[idx]:
            st.markdown(f"### {route_info['icon']} {route_info['name']}")
            st.write(route_info['description'])
            st.write(f"**Modes**: {', '.join([mm['type'] for mm in route_info['modes']])}")
            st.write(f"**CO2**: {route_info['co2']:.1f} tons")

            total_cost_mult = sum([TRANSPORT_MODES[mm['type']]['cost_multiplier'] * mm['percentage']/100 for mm in route_info['modes']])
            total_time_mult = sum([TRANSPORT_MODES[mm['type']]['time_multiplier'] * mm['percentage']/100 for mm in route_info['modes']])

            est_hours = max(1, int(total_distance / (vehicle_speed / total_time_mult) * 24))
            st.write(f"**Est. Time**: {est_hours} hours")
            st.write(f"**Cost Index**: {total_cost_mult:.2f}x")

    st.subheader("AI-Powered Strategy Recommendation")
    strategies = generate_default_strategies_domestic(
        origin_input, destination_input, total_distance, cargo_type, dangerous_goods, vehicle_speed
    )

    col_strat_a, col_strat_b, col_strat_c = st.columns(3)

    with col_strat_a:
        strat_key = list(strategies.keys())[0]
        strat = strategies[strat_key]
        st.markdown(f"### {strat.get('name', 'Strategy A')}")
        st.write(strat.get('description', ''))
        st.write(f"**Time**: {strat.get('time_days', 1)} days")
        st.write(f"**Cost**: {strat.get('cost_multiplier', 1.5):.2f}x")
        st.write(f"**Risk**: {strat.get('risk_score', 30)}/100")
        st.write("**Modes**: " + ", ".join(strat.get('modes', [])))
        if st.button("Select Strategy A", key="btn_domestic_a"):
            st.session_state.selected_strategy = strat_key
            st.rerun()

    with col_strat_b:
        strat_key = list(strategies.keys())[1]
        strat = strategies.get(strat_key, {})
        st.markdown(f"### {strat.get('name', 'Strategy B')}")
        st.write(strat.get('description', ''))
        st.write(f"**Time**: {strat.get('time_days', 2)} days")
        st.write(f"**Cost**: {strat.get('cost_multiplier', 1.0):.2f}x")
        st.write(f"**Risk**: {strat.get('risk_score', 45)}/100")
        st.write("**Modes**: " + ", ".join(strat.get('modes', [])))
        if st.button("Select Strategy B", key="btn_domestic_b"):
            st.session_state.selected_strategy = strat_key
            st.rerun()

    with col_strat_c:
        strat_key = list(strategies.keys())[2]
        strat = strategies.get(strat_key, {})
        st.markdown(f"### {strat.get('name', 'Strategy C')}")
        st.write(strat.get('description', ''))
        st.write(f"**Time**: {strat.get('time_days', 3)} days")
        st.write(f"**Cost**: {strat.get('cost_multiplier', 0.7):.2f}x")
        st.write(f"**Risk**: {strat.get('risk_score', 55)}/100")
        st.write("**Modes**: " + ", ".join(strat.get('modes', [])))
        if st.button("Select Strategy C", key="btn_domestic_c"):
            st.session_state.selected_strategy = strat_key
            st.rerun()

    render_strategy_comparison(strategies, total_distance, cargo_weight, container_40, container_20, dangerous_goods, key_prefix="domestic")
    render_route_comparison(domestic_coords_lookup, origin_input, key_prefix="domestic", is_domestic=True)
    render_social_benefit_section(total_distance)
    render_rail_logistics_platform()
    render_repositioning_section(destination_input)

    st.subheader("Strategy Scenario Simulation")
    selected_strat = strategies.get(st.session_state.selected_strategy, strategies.get(list(strategies.keys())[1], {}))

    if st.button("Run Detailed Simulation", key="sim_domestic_btn"):
        st.session_state.run_simulation = True

    sim_result = None
    if st.session_state.run_simulation:
        with st.spinner("Simulating..."):
            sim_result = run_scenario_simulation(
                selected_strat, total_distance, cargo_weight, container_40, container_20, dangerous_goods, 0
            )
        db_save_simulation_history("국내운송", origin_input, destination_input, selected_strat.get("name", ""), sim_result)

        cur = st.session_state.currency
        col_sim1, col_sim2, col_sim3 = st.columns(3)
        with col_sim1:
            st.metric("Expected Cost", format_currency(sim_result['cost_mean'], cur),
                       delta=f"±{format_currency(sim_result['cost_std'], cur)}")
        with col_sim2:
            st.metric("Expected Delivery", f"{sim_result['time_mean']:.1f} days", delta=f"±{sim_result['time_std']:.1f} days")
        with col_sim3:
            st.metric("Risk Score", f"{sim_result['risk_mean']:.0f}/100", delta=f"±{sim_result['risk_std']:.0f}")

        render_simulation_charts(sim_result)
        check_and_alert_risk(sim_result, strategies, st.session_state.selected_strategy, f"{origin_input} → {destination_input}", key_prefix="domestic")

        if st.button("진행중인 배송으로 등록", key="domestic_register_shipment"):
            db_save_active_shipment(f"{origin_input} → {destination_input}", "국내운송", origin_input, destination_input,
                                     selected_strat.get("name", ""), sim_result)
            st.success("진행중인 배송으로 등록했습니다. 홈 대시보드에서 확인하세요.")

    render_accept_reject_section(sim_result, key_prefix="domestic")

    with st.expander(t("calculator_title")):
        est_cargo_value = st.number_input("화물가액 (USD, 개산)", min_value=0.0, value=50000.0, step=1000.0, key="domestic_cargo_value")
        risk_for_insurance = selected_strat.get("risk_score", 45)
        premium, rate = estimate_insurance_premium(est_cargo_value, risk_for_insurance)
        st.metric("개산 보험료", format_currency(premium, st.session_state.currency), delta=f"요율 {rate*100:.2f}%")
        st.caption("보험료 = 화물가액 × (기본요율 0.15% + 리스크스코어 비례 가산 최대 1.35%p)")

        st.divider()
        st.write(t("forwarder_header"))
        render_forwarder_reputation()

    domestic_route_info = {
        "구분": "국내운송", "출발지": origin_input, "도착지": destination_input,
        "경유지": ", ".join(transit_inputs) if transit_inputs else "없음",
        "총거리(km)": f"{total_distance:.0f}", "화물중량(톤)": cargo_weight,
        "화물종류": cargo_type, "위험물여부": "예" if dangerous_goods else "아니오",
    }
    render_report_and_share_section("ATLAS AI 국내운송 리포트", domestic_route_info, selected_strat, sim_result, key_prefix="domestic")

    render_favorites_manager({
        "type": "국내운송", "origin": origin_input, "destination": destination_input,
        "transit": ", ".join(transit_inputs), "distance_km": round(total_distance, 0),
    }, key_prefix="domestic")

with tab2:
    st.header(t("intl_header"))

    render_ai_dashboard()

    st.divider()

    if st.session_state.logistics_insights is None:
        with st.spinner("AI가 물류 뉴스를 학습 중..."):
            news_items = fetch_logistics_news()
            st.session_state.logistics_insights = news_items
    else:
        news_items = st.session_state.logistics_insights

    with st.expander(t("ai_assistant_title"), expanded=False):
        if not GEMINI_AVAILABLE:
            st.caption("`pip install google-generativeai` 설치 후 사용 가능합니다." if st.session_state.nav_language == "ko" else "Available after `pip install google-generativeai`.")
        elif not st.session_state.gemini_api_key:
            st.caption("사이드바 '전역 설정 / API 연동'에서 Gemini API Key를 입력하면 활성화됩니다." if st.session_state.nav_language == "ko" else "Enter your Gemini API Key in the sidebar 'Global Settings' to enable this.")
        else:
            user_question = st.text_input(t("ai_question_label"), placeholder="예: 다음 주 부산→로테르담 항로 리스크는?" if st.session_state.nav_language == "ko" else "e.g. What's the risk on the Busan→Rotterdam route next week?", key="ai_question_intl")
            if st.button(t("ai_ask_btn"), key="ai_ask_btn_intl") and user_question:
                with st.spinner("Gemini가 답변을 생성 중..." if st.session_state.nav_language == "ko" else "Gemini is generating a response..."):
                    answer = ask_gemini(user_question, st.session_state.gemini_api_key)
                st.session_state.ai_chat_history.append((user_question, answer))
            for q, a in reversed(st.session_state.ai_chat_history[-5:]):
                st.markdown(f"**Q. {q}**")
                st.write(a)
                st.divider()

        if st.button(t("ai_briefing_btn"), key="ai_briefing_btn_intl"):
            with st.spinner("최신 뉴스를 분석 중..."):
                st.session_state.ai_risk_briefing = generate_ai_risk_briefing(news_items, st.session_state.gemini_api_key)
        if st.session_state.ai_risk_briefing:
            st.info(st.session_state.ai_risk_briefing)

    with st.sidebar.expander(t("sidebar_intl_settings"), expanded=False):
        incoterms = st.selectbox("Incoterms", ["EXW","FCA","FAS","FOB","CFR","CIF","CPT","CIP","DAP","DPU","DDP"], index=4, key="intl_incoterms")
        cargo_type_intl = st.selectbox("Cargo Type", ["Automotive Parts", "Semiconductors", "Batteries", "Raw Materials", "General Cargo", "Electronics"], key="intl_cargo")

        st.divider()
        st.write("**Cargo Details**")

        is_bulk_cargo_intl = st.checkbox("Bulk Cargo", value=False, key="intl_bulk")

        if is_bulk_cargo_intl:
            st.write("**Bulk Cargo Settings**")
            bulk_type_intl = st.selectbox("Bulk Cargo Type", BULK_CARGO_TYPES, key="intl_bulk_type")

            col_bulk1, col_bulk2 = st.columns(2)
            with col_bulk1:
                bulk_density_intl = st.number_input("Density (ton/m³)", min_value=0.1, value=1.5, step=0.1, key="intl_bulk_density")
            with col_bulk2:
                bulk_hazard_intl = st.selectbox("Hazard Level", ["Low", "Medium", "High"], key="intl_bulk_hazard")

            bulk_stowage_intl = st.selectbox("Stowage", ["Special", "Ventilated", "Certified", "Standard"], key="intl_bulk_stowage")
            bulk_quantity_intl = st.number_input("Quantity (tons)", min_value=0.0, value=100.0, step=10.0, key="intl_bulk_qty")
            bulk_volume_intl = bulk_quantity_intl / bulk_density_intl if bulk_density_intl > 0 else 0
            st.write(f"**Estimated Volume: {bulk_volume_intl:.2f} m³**")

            cargo_weight_intl = bulk_quantity_intl
            container_40_intl = 0
            container_20_intl = 0
        else:
            commodity_intl = st.text_area("Commodity Description", value="", height=50, placeholder="화물 설명 입력", key="intl_commodity")
            cargo_weight_intl = st.number_input("Total Weight (tons)", min_value=0.0, value=10.0, step=0.5, key="intl_weight")
            container_40_intl = st.number_input('40ft Containers', min_value=0, max_value=100, value=1, key="intl_40ft")
            container_20_intl = st.number_input('20ft Containers', min_value=0, max_value=100, value=0, key="intl_20ft")

        dangerous_goods_intl = st.checkbox("Dangerous Goods", value=False, key="intl_danger")

        st.divider()
        st.write("**Schedule**")
        dep_date_intl = st.date_input("Departure Date", value=date.today(), key="intl_dep_date")
        dep_time_intl = st.time_input("Departure Time", value=dtime(hour=9, minute=0), key="intl_dep_time")
        desired_arrival_intl = st.date_input("Desired Arrival", value=date.today() + timedelta(days=30), key="intl_arrival")

        st.divider()
        st.write("**Route Selection**")
        origin_intl = st.selectbox("Origin Port", list(PORTS_DB.keys()), index=0, key="intl_origin")

        transit_port_options = [p for p in PORTS_DB.keys() if p != origin_intl]
        transit_ports = st.multiselect("Transit Ports (선택)", transit_port_options, key="intl_transit_ports")

        destination_intl = st.selectbox("Destination Port", [p for p in PORTS_DB.keys() if p != origin_intl], key="intl_destination")

        st.divider()
        st.write("**Shipment Mode (Door-to-Door)**")
        shipment_mode_intl = st.radio(
            "운송 형태", ["Port-to-Port", "Door-to-Door (라스트마일 포함)"], key="intl_shipment_mode"
        )
        if shipment_mode_intl.startswith("Door-to-Door"):
            domestic_door_origin = st.selectbox(
                "국내 출발지 (공장/창고)", list(DOMESTIC_TERMINALS.keys()), key="intl_door_origin"
            )
            final_destination_text = st.text_input(
                "최종 도착지 (자유 입력, 예: Rotterdam 물류센터)", key="intl_door_dest"
            )
        else:
            domestic_door_origin = None
            final_destination_text = ""

        st.divider()
        st.write("**Cargo Consolidation**")
        cargo_consolidation_intl = st.radio(
            "화물 적재 방식", ["FCL (풀컨테이너)", "LCL (혼적/개별화주)"], key="intl_consolidation"
        )

        st.divider()
        st.write("**Geopolitical Risks**")
        selected_risks = st.multiselect(
            "Select Affected Areas",
            list(GEOPOLITICAL_RISKS.keys()),
            key="intl_risks"
        )

        st.divider()
        st.write("**Transportation Modes**")
        use_air_intl = st.checkbox("Air Transport", value=False, key="intl_air")
        use_sea_intl = st.checkbox("Sea Transport", value=True, key="intl_sea")
        use_rail_intl = st.checkbox("Rail Transport", value=False, key="intl_rail")
        use_road_intl = st.checkbox("Road Transport", value=False, key="intl_road")

        st.divider()
        st.write("**Settings**")
        vessel_speed_intl = st.slider("Average Speed (knots)", 8, 25, 12, key="intl_speed")
        show_map_intl = st.checkbox("Show Interactive Map", value=True, key="intl_show_map")

    origin_coord_intl = PORTS_DB[origin_intl][:2]
    destination_coord_intl = PORTS_DB[destination_intl][:2]
    transit_port_coords = [PORTS_DB[p][:2] for p in transit_ports] if transit_ports else []

    route_distance_intl = haversine_km(origin_coord_intl[0], origin_coord_intl[1],
                                       transit_port_coords[0][0] if transit_port_coords else destination_coord_intl[0],
                                       transit_port_coords[0][1] if transit_port_coords else destination_coord_intl[1])

    for i in range(len(transit_port_coords) - 1):
        route_distance_intl += haversine_km(transit_port_coords[i][0], transit_port_coords[i][1],
                                           transit_port_coords[i+1][0], transit_port_coords[i+1][1])

    if transit_port_coords:
        route_distance_intl += haversine_km(transit_port_coords[-1][0], transit_port_coords[-1][1],
                                           destination_coord_intl[0], destination_coord_intl[1])

    adjusted_distance, risk_multiplier, risk_desc = apply_geopolitical_risk(route_distance_intl, selected_risks)

    ordered_names_intl = [origin_intl] + transit_ports + [destination_intl]
    intl_coords_lookup = {k: v[:2] for k, v in PORTS_DB.items()}

    if st.session_state.ais_vessels_intl is None:
        real_vessels_intl = None
        if st.session_state.aishub_username:
            lats = [intl_coords_lookup[n][1] for n in ordered_names_intl]
            lons = [intl_coords_lookup[n][0] for n in ordered_names_intl]
            real_vessels_intl = fetch_real_ais_vessels(
                min(lats) - 5, max(lats) + 5, min(lons) - 5, max(lons) + 5, st.session_state.aishub_username
            )
        st.session_state.ais_vessels_intl = real_vessels_intl or generate_realistic_ais_vessels(150)
        st.session_state.ais_source_intl = "AISHub (실데이터)" if real_vessels_intl else "시뮬레이션"

    typhoons_intl = fetch_typhoon_paths(st.session_state.kma_api_key) if st.session_state.kma_api_key else []

    if show_map_intl:
        st.subheader("Real-time Global Maritime Map (MarineTraffic AIS)")
        vessel_source_label_intl = st.session_state.get("ais_source_intl", "시뮬레이션")
        st.markdown(f"**Active Vessels: {len(st.session_state.ais_vessels_intl)}+ ({vessel_source_label_intl}) | Ports: 24 | Risk Zones | ━━ Routes**")

        leg_options_intl = ["전체 경로"] + [f"{ordered_names_intl[i]} → {ordered_names_intl[i+1]}"
                                              for i in range(len(ordered_names_intl) - 1)]
        st.session_state.map_zoom_leg_intl = st.selectbox(
            "구간 확대", leg_options_intl, key="intl_leg_zoom",
            index=leg_options_intl.index(st.session_state.map_zoom_leg_intl)
            if st.session_state.map_zoom_leg_intl in leg_options_intl else 0
        )

        default_center_intl = [-10.5, -51.3]
        map_center_intl, map_zoom_intl = get_map_view_for_leg(
            st.session_state.map_zoom_leg_intl, ordered_names_intl, intl_coords_lookup, default_center_intl, 2
        )

        m_intl = folium.Map(
            location=map_center_intl,
            zoom_start=map_zoom_intl,
            tiles="CartoDB positron"
        )

        MiniMap().add_to(m_intl)
        Fullscreen().add_to(m_intl)

        render_risk_zones_circle(m_intl, selected_risks)

        route_waypoints = [[PORTS_DB[p][1], PORTS_DB[p][0]] for p in transit_ports]
        base_route = [[origin_coord_intl[1], origin_coord_intl[0]]] + route_waypoints + [[destination_coord_intl[1], destination_coord_intl[0]]]

        folium.PolyLine(
            locations=base_route,
            color="#0077BE",
            weight=3,
            opacity=0.8,
            popup="Base Route",
            tooltip="Base Route",
            dash_array=None
        ).add_to(m_intl)

        ai_route_intl = create_ai_route_with_waypoints(origin_coord_intl,
                                                       transit_port_coords,
                                                       destination_coord_intl,
                                                       num_points=15, offset_km=20)

        folium.PolyLine(
            locations=ai_route_intl,
            color="#FF7F0E",
            weight=3,
            opacity=0.7,
            popup="AI Optimized Route",
            tooltip="AI Route (Optimized)",
            dash_array="5, 5"
        ).add_to(m_intl)

        for port_name, port_info in PORTS_DB.items():
            lon, lat = port_info[:2]

            if port_name == origin_intl:
                color = "#1f77b4"
            elif port_name == destination_intl:
                color = "#ff7f0e"
            elif port_name in transit_ports:
                color = "#2ca02c"
            else:
                color = "#95a5a6"

            folium.CircleMarker(
                location=[lat, lon],
                radius=5,
                color=color,
                fill=True,
                fillColor=color,
                fillOpacity=0.8,
                weight=1.5,
                popup=f"<b>{port_name}</b>",
                tooltip=port_name
            ).add_to(m_intl)

        render_ais_vessels_on_map(m_intl, st.session_state.ais_vessels_intl)
        render_typhoon_paths_on_map(m_intl, typhoons_intl)

        render_folium_map(m_intl, height=750, map_key="intl_map")

        if typhoons_intl:
            st.warning(f"활성 태풍 {len(typhoons_intl)}건 감지 - 해상 경로 리스크 참고")

        col_risk1, col_risk2 = st.columns([3, 1])
        with col_risk2:
            st.subheader("리스크 분석")
            if selected_risks:
                st.warning(f"**Risk Multiplier: {risk_multiplier:.2f}x**")
                for risk in risk_desc:
                    st.write(f"• {risk}")
            else:
                st.success("No Selected Risks")

        cong_col1, cong_col2 = cols_adaptive(2)
        with cong_col1:
            cg_o = fetch_port_congestion_index(origin_intl)
            st.metric(f"{origin_intl.split(' (')[0]} 혼잡도", f"{cg_o['congestion_score']}/100", delta=f"대기 {cg_o['waiting_days']}일")
        with cong_col2:
            cg_d = fetch_port_congestion_index(destination_intl)
            st.metric(f"{destination_intl.split(' (')[0]} 혼잡도", f"{cg_d['congestion_score']}/100", delta=f"대기 {cg_d['waiting_days']}일")

    render_ais_fleet_statistics(st.session_state.ais_vessels_intl, coverage_label="Global")

    st.subheader("Global Port Statistics")
    col_stat1, col_stat2 = st.columns(2)
    with col_stat1:
        st.metric("Total Ports", len(PORTS_DB))
    with col_stat2:
        st.metric("Route Distance", f"{route_distance_intl:,.0f} km")

    st.subheader("Multi-Modal Transportation Routes")
    multimodal_routes_intl = generate_multimodal_routes_international(adjusted_distance)

    cols_intl = st.columns(len(multimodal_routes_intl))

    for idx, (route_key, route_info) in enumerate(multimodal_routes_intl.items()):
        with cols_intl[idx]:
            st.markdown(f"### {route_info['icon']} {route_info['name']}")
            st.write(route_info['description'])
            st.write(f"**Modes**: {', '.join([mm['type'] for mm in route_info['modes']])}")
            st.write(f"**CO2**: {route_info['co2']:.1f} tons")

            total_cost_mult_intl = sum([TRANSPORT_MODES[mm['type']]['cost_multiplier'] * mm['percentage']/100 for mm in route_info['modes']])
            total_time_mult_intl = sum([TRANSPORT_MODES[mm['type']]['time_multiplier'] * mm['percentage']/100 for mm in route_info['modes']])

            est_days_intl = max(3, int(adjusted_distance / (vessel_speed_intl * 1.852 * 24) * total_time_mult_intl))
            st.write(f"**Est. Time**: {est_days_intl} days")
            st.write(f"**Cost Index**: {total_cost_mult_intl:.2f}x")

    st.subheader("AI-Powered Strategy Recommendation")
    strategies_intl = generate_default_strategies_international(
        origin_intl, destination_intl, adjusted_distance, cargo_type_intl, dangerous_goods_intl, vessel_speed_intl
    )

    col_strat_a, col_strat_b, col_strat_c = st.columns(3)

    with col_strat_a:
        strat_key = list(strategies_intl.keys())[0]
        strat = strategies_intl[strat_key]
        st.markdown(f"### {strat.get('name', 'Strategy A')}")
        st.write(strat.get('description', ''))
        st.write(f"**Time**: {strat.get('time_days', 5)} days")
        st.write(f"**Cost**: {strat.get('cost_multiplier', 1.35):.2f}x")
        st.write(f"**Risk**: {strat.get('risk_score', 45)}/100")
        st.write("**Modes**: " + ", ".join(strat.get('modes', [])))
        if st.button("Select Strategy A", key="btn_intl_a"):
            st.session_state.selected_strategy = strat_key
            st.rerun()

    with col_strat_b:
        strat_key = list(strategies_intl.keys())[1]
        strat = strategies_intl.get(strat_key, {})
        st.markdown(f"### {strat.get('name', 'Strategy B')}")
        st.write(strat.get('description', ''))
        st.write(f"**Time**: {strat.get('time_days', 15)} days")
        st.write(f"**Cost**: {strat.get('cost_multiplier', 1.0):.2f}x")
        st.write(f"**Risk**: {strat.get('risk_score', 60)}/100")
        st.write("**Modes**: " + ", ".join(strat.get('modes', [])))
        if st.button("Select Strategy B", key="btn_intl_b"):
            st.session_state.selected_strategy = strat_key
            st.rerun()

    with col_strat_c:
        strat_key = list(strategies_intl.keys())[2]
        strat = strategies_intl.get(strat_key, {})
        st.markdown(f"### {strat.get('name', 'Strategy C')}")
        st.write(strat.get('description', ''))
        st.write(f"**Time**: {strat.get('time_days', 25)} days")
        st.write(f"**Cost**: {strat.get('cost_multiplier', 0.78):.2f}x")
        st.write(f"**Risk**: {strat.get('risk_score', 75)}/100")
        st.write("**Modes**: " + ", ".join(strat.get('modes', [])))
        if st.button("Select Strategy C", key="btn_intl_c"):
            st.session_state.selected_strategy = strat_key
            st.rerun()

    render_strategy_comparison(strategies_intl, adjusted_distance, cargo_weight_intl, container_40_intl, container_20_intl, dangerous_goods_intl, key_prefix="intl")
    render_route_comparison(intl_coords_lookup, origin_intl, key_prefix="intl", is_domestic=False)
    st.caption("전략 비교 및 경로 비교는 Port-to-Port 기준 개산치이며, Door-to-Door 라스트마일/LCL 보정은 아래 Strategy Scenario Simulation에만 반영됩니다.")

    st.subheader("Strategy Scenario Simulation")
    selected_strat_intl = strategies_intl.get(st.session_state.selected_strategy, strategies_intl.get(list(strategies_intl.keys())[1], {}))

    if st.button("Run Detailed Simulation", key="sim_intl_btn"):
        st.session_state.run_simulation = True

    sim_result_intl = None
    inland_port_name, inland_dist_km, inland_time_days = None, 0.0, 0.0
    if shipment_mode_intl.startswith("Door-to-Door") and domestic_door_origin:
        inland_port_name, inland_dist_km = find_nearest_korean_port(domestic_door_origin)
        inland_time_days = inland_dist_km / DOMESTIC_TRUCK_SPEED_KMH / 24

    if st.session_state.run_simulation:
        with st.spinner("Simulating..."):
            sim_result_intl = run_scenario_simulation(
                selected_strat_intl, adjusted_distance, cargo_weight_intl, container_40_intl, container_20_intl, dangerous_goods_intl, 0
            )

            if cargo_consolidation_intl.startswith("LCL"):
                sim_result_intl = apply_shipment_mode_adjustments(
                    sim_result_intl, extra_days=LCL_CONSOLIDATION_EXTRA_DAYS, cost_multiplier=LCL_COST_MULTIPLIER
                )

            if shipment_mode_intl.startswith("Door-to-Door") and domestic_door_origin:
                sim_result_intl = apply_shipment_mode_adjustments(
                    sim_result_intl,
                    extra_days=inland_time_days + DOOR_TO_DOOR_LASTMILE_DAYS,
                    cost_multiplier=(1 + DOOR_TO_DOOR_LASTMILE_COST_RATE)
                )

        db_save_simulation_history("국제운송", origin_intl, destination_intl, selected_strat_intl.get("name", ""), sim_result_intl)

        if shipment_mode_intl.startswith("Door-to-Door") and domestic_door_origin:
            st.info(f"Door-to-Door: {domestic_door_origin} -> 최근접 국내항 {inland_port_name}"
                    f"(내륙 {inland_dist_km:.0f}km, 약 {inland_time_days:.1f}일) + 도착지 라스트마일 버퍼 {DOOR_TO_DOOR_LASTMILE_DAYS}일, "
                    f"운임 {DOOR_TO_DOOR_LASTMILE_COST_RATE*100:.0f}% 추가 반영됨"
                    + (f" / 최종 도착지: {final_destination_text}" if final_destination_text else ""))
        if cargo_consolidation_intl.startswith("LCL"):
            st.caption(f"LCL(혼적) 반영: 통합 대기 +{LCL_CONSOLIDATION_EXTRA_DAYS:.0f}일, 운임 {LCL_COST_MULTIPLIER:.2f}x")

        cur_intl = st.session_state.currency
        col_sim1, col_sim2, col_sim3 = st.columns(3)
        with col_sim1:
            st.metric("Expected Cost", format_currency(sim_result_intl['cost_mean'], cur_intl),
                       delta=f"±{format_currency(sim_result_intl['cost_std'], cur_intl)}")
        with col_sim2:
            st.metric("Expected Delivery", f"{sim_result_intl['time_mean']:.1f} days", delta=f"±{sim_result_intl['time_std']:.1f} days")
        with col_sim3:
            st.metric("Risk Score", f"{sim_result_intl['risk_mean']:.0f}/100", delta=f"±{sim_result_intl['risk_std']:.0f}")

        render_simulation_charts(sim_result_intl)
        check_and_alert_risk(sim_result_intl, strategies_intl, st.session_state.selected_strategy, f"{origin_intl} → {destination_intl}", key_prefix="intl")
        render_delay_buffer_warning(sim_result_intl)

        if st.button("진행중인 배송으로 등록", key="intl_register_shipment"):
            db_save_active_shipment(f"{origin_intl} → {destination_intl}", "국제운송", origin_intl, destination_intl,
                                     selected_strat_intl.get("name", ""), sim_result_intl)
            st.success("진행중인 배송으로 등록했습니다. 홈 대시보드에서 확인하세요.")

    render_accept_reject_section(sim_result_intl, key_prefix="intl")

    st.subheader("Route Statistics (Risk Reflected)")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Base Distance", f"{route_distance_intl:,.0f} km")
    with col2:
        st.metric("Adjusted Distance", f"{adjusted_distance:,.0f} km")
    with col3:
        est_days = max(3, int(adjusted_distance / (vessel_speed_intl * 1.852 * 24)))
        st.metric("Est. Duration", f"{est_days} days")
    with col4:
        st.metric("Cost Multiplier", f"{risk_multiplier:.2f}x")

    with st.expander(t("calculator_title")):
        clr = calculate_customs_clearance_days(incoterms, cargo_type_intl, dangerous_goods_intl)
        cc1, cc2, cc3, cc4 = cols_adaptive(4)
        with cc1:
            st.metric("수출통관", f"{clr['export_days']}일")
        with cc2:
            st.metric("수입통관", f"{clr['import_days']}일")
        with cc3:
            st.metric("추가지연", f"{clr['extra_days']}일")
        with cc4:
            st.metric("총 통관일수", f"{clr['total_days']}일")
        st.caption(f"Incoterms {incoterms}: {clr['note']}")

        st.divider()
        dest_country = PORTS_DB[destination_intl][2]
        eu_countries = {"Netherlands", "Germany", "Belgium", "France", "Spain", "Greece"}
        est_co2 = multimodal_routes_intl.get("FULL_SEA", {}).get("co2", 0.2 * (adjusted_distance / 1000))
        if dest_country in eu_countries:
            cbam_cost = calculate_cbam_cost(est_co2)
            st.metric("CBAM 개산 비용 (EU 도착)", format_currency(cbam_cost, st.session_state.currency),
                       delta=f"CO2 {est_co2:.1f}톤 × €{EU_CARBON_PRICE_EUR_PER_TON}/톤")
            st.caption("탄소국경조정제도(CBAM) 참고용 근사치이며 실제 부과액은 품목별 벤치마크에 따라 달라집니다.")
        else:
            cbam_cost = 0.0
            st.caption(f"도착지({dest_country})는 EU 역외라 CBAM 대상이 아닙니다. (참고 CO2 배출량: {est_co2:.1f}톤)")

        st.divider()
        st.write("**HS코드 관세 계산기**")
        hs_code_input = st.text_input("HS코드 (앞 4자리, 예: 8708 자동차부품)", value="8708", key="intl_hs_code")
        tariff_info = fetch_customs_tariff_rate(hs_code_input, st.session_state.customs_api_key)
        est_cargo_value_intl = st.number_input("화물가액 (USD, 개산)", min_value=0.0, value=200000.0, step=5000.0, key="intl_cargo_value")
        if tariff_info.get("rate") is not None:
            duty = est_cargo_value_intl * (tariff_info["rate"] / 100)
            st.metric(f"개산 관세 ({tariff_info.get('desc', hs_code_input)})", format_currency(duty, st.session_state.currency),
                       delta=f"세율 {tariff_info['rate']}% · {tariff_info['source']}")
        else:
            duty = 0.0
            st.info(f"관세청 API 응답: {tariff_info.get('raw_response', '(파싱 필요)')}")
        st.caption("실서비스 전환 시 관세청 unipass 오픈API의 정확한 엔드포인트/파라미터 확인이 필요합니다.")

        st.divider()
        risk_for_insurance_intl = selected_strat_intl.get("risk_score", 60)
        premium_intl, rate_intl = estimate_insurance_premium(est_cargo_value_intl, risk_for_insurance_intl)
        st.metric("개산 보험료", format_currency(premium_intl, st.session_state.currency), delta=f"요율 {rate_intl*100:.2f}%")

        st.divider()
        freight_cost_for_tco = sim_result_intl["cost_mean"] if sim_result_intl else selected_strat_intl.get("cost_multiplier", 1.0) * 5000
        tco = calculate_tco(freight_cost_for_tco, duty, cbam_cost, premium_intl)
        st.metric("총소요비용 (TCO 개산)", format_currency(tco, st.session_state.currency),
                   delta="운임 + 관세 + CBAM + 보험료")

        st.divider()
        st.write(t("forwarder_header"))
        render_forwarder_reputation()

    intl_route_info = {
        "구분": "국제운송", "Incoterms": incoterms, "출발항": origin_intl, "도착항": destination_intl,
        "경유항": ", ".join(transit_ports) if transit_ports else "없음",
        "기본거리(km)": f"{route_distance_intl:,.0f}", "리스크반영거리(km)": f"{adjusted_distance:,.0f}",
        "화물종류": cargo_type_intl, "위험물여부": "예" if dangerous_goods_intl else "아니오",
        "선택 리스크": ", ".join(selected_risks) if selected_risks else "없음",
        "운송형태": shipment_mode_intl, "화물적재방식": cargo_consolidation_intl,
        "국내출발지": domestic_door_origin if domestic_door_origin else "-",
        "최종도착지": final_destination_text if final_destination_text else "-",
    }
    render_report_and_share_section("ATLAS AI 국제운송 리포트", intl_route_info, selected_strat_intl, sim_result_intl, key_prefix="intl")

    render_favorites_manager({
        "type": "국제운송", "origin": origin_intl, "destination": destination_intl,
        "transit": ", ".join(transit_ports), "distance_km": round(adjusted_distance, 0),
    }, key_prefix="intl")

st.markdown("---")
st.info("ATLAS AI v9.1 - Real-time AIS/FX/Weather Integration | AI Assistant | Persistent DB (Favorites/History/Shipments) | Strategy & Route Comparison | Risk Auto-Alert & Reroute | HS Tariff/TCO Calculator | KO/EN | Multi-Stop Routes | Monte Carlo Simulation")
